"""
Multi-Agente LangGraph — Orquestrador + Sub-agente Analista de Dados
========================================================================
Arquitetura:

    Usuário
       ↓ query
    Orquestrador (StateGraph)
       ├─ get_current_datetime()     → responde data/hora direto
       └─ consultar_analista(detalhes) → delega ao sub-agente
              └─ Sub-agente analista (StateGraph)
                    ├─ read_skill(filename)              → lê instruções do arquivo .md
                    ├─ chamar_api(url, params)            → chama a API REST
                    └─ executar_codigo_pandas(codigo, chave) → sandbox Python/Pandas

Fluxo do sub-agente por análise:
  1. read_skill → descobre a API e a estrutura do JSON
  2. chamar_api(url, chave) → busca dados e salva JSON em api_data[chave] (sem expor ao LLM)
  3. executar_codigo_pandas(codigo, chave) → cria df, analisa, retorna resultado formatado
"""

import contextvars
import io
import logging
import os
import queue as _queue_module
import sqlite3
import sys
import textwrap
import threading
import traceback as tb
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import requests as http_requests
from langchain_core.messages import BaseMessage, HumanMessage, SystemMessage
from langchain_core.tools import tool
from langgraph.graph import END, START, StateGraph
from langgraph.graph.message import add_messages
from langgraph.prebuilt import ToolNode
from typing_extensions import Annotated, TypedDict

import chart_store
from scheduler.tools import (
    schedule_task,
    get_task_code,
    test_task_code,
    save_task_code,
)
from scheduler.widget_tools import (
    add_chart_to_dashboard,
    test_widget_code,
    list_dashboard_widgets,
    delete_dashboard_widget,
    get_widget_code,
    update_widget,
)

logger = logging.getLogger(__name__)

from db import DB_PATH

SKILLS_FOLDER = Path(__file__).parent / "skills"
SKILL_HEADER_LINES = 4

_orchestrator_graph = None
_sub_agent_graph = None
_scheduler_agent_graph = None
_scheduling_graph = None
_checkpointer = None

# Propaga o session_id através das fronteiras de thread do ToolNode.
# contextvars.ContextVar se propaga para threads criadas por ThreadPoolExecutor,
# que é o que o ToolNode do LangGraph usa internamente.
_current_session: contextvars.ContextVar[str] = contextvars.ContextVar(
    "mfg_session", default="default"
)
_sandbox_mode: contextvars.ContextVar[bool] = contextvars.ContextVar(
    "sandbox_mode", default=False
)
_sandbox_task_id: contextvars.ContextVar[str] = contextvars.ContextVar(
    "sandbox_task_id", default=""
)

# Snapshot do dashboard enviado pelo browser antes de cada mensagem.
# Chave: session_id, Valor: dict com charts e kpis.
_chart_snapshots: dict[str, dict] = {}
_cs_lock = threading.Lock()


def set_chart_snapshot(session_id: str, data: dict) -> None:
    """Chamado pelo endpoint POST /chart-snapshot para armazenar o estado atual do dashboard."""
    with _cs_lock:
        _chart_snapshots[session_id] = data


def _tlog(tool: str, event: str, **kwargs) -> None:
    """Loga chamadas e saídas de tools quando DEBUG_TOOLS=1."""
    if os.getenv("DEBUG_TOOLS", "0").strip() != "1":
        return
    lines = [f"\n{'='*60}", f"[TOOL] {tool} | {event}"]
    truncate = int(os.getenv("DEBUG_TOOLS_TRUNCATE", "0"))
    for k, v in kwargs.items():
        text = str(v)
        if truncate > 0 and len(text) > truncate:
            text = text[:truncate] + " …(truncado)"
        lines.append(f"  {k}: {text}")
    lines.append("="*60)
    logger.info("\n".join(lines))


# ── Namespace persistente por sessão (estilo Jupyter) ─────────────────────────
# Cada sessão tem um dict de execução Python que persiste entre tool calls e
# entre invocações do sub-agente. O LLM recebe um resumo das variáveis ao final
# de cada tool para saber o que já está disponível para análises incrementais.

import matplotlib
matplotlib.use("Agg")  # backend sem display — obrigatório em servidor
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from scipy import stats as scipy_stats

_BUILTINS_SKIP = {"pd", "np", "plt", "stats", "__builtins__", "__doc__", "__name__",
                  "__package__", "__spec__", "__loader__", "__import__", "__build_class__"}

_namespaces: dict = {}       # session → namespace dict
_ns_last_access: dict = {}   # session → datetime do último acesso
_ns_lock = threading.Lock()
_NS_TTL_SECONDS = 600        # 10 minutos de inatividade


def _ns_get() -> dict:
    session = _current_session.get()
    with _ns_lock:
        if session not in _namespaces:
            _namespaces[session] = {"pd": pd, "np": np, "plt": plt, "stats": scipy_stats}
        _ns_last_access[session] = datetime.now()
        return _namespaces[session]


def _ns_context_for_agent(ns: dict) -> tuple[str, list]:
    """Retorna (vars_summary, script_history) para injetar no sub-agente como contexto."""
    _SKIP = _BUILTINS_SKIP | {"stats", "_current_query", "_script_history"}
    lines = []
    for k, v in ns.items():
        if k.startswith("_") or k in _SKIP:
            continue
        if isinstance(v, pd.DataFrame):
            lines.append(f"  - '{k}': DataFrame {len(v)} linhas, colunas: {list(v.columns)}")
        elif isinstance(v, pd.Series):
            lines.append(f"  - '{k}': Series {len(v)} itens")
    history = ns.get("_script_history", [])
    return "\n".join(lines), history


async def ns_cleanup_loop():
    """Loop assíncrono que expira namespaces inativos. Registrar via asyncio.create_task no startup."""
    import asyncio
    while True:
        await asyncio.sleep(60)
        cutoff = datetime.now() - timedelta(seconds=_NS_TTL_SECONDS)
        with _ns_lock:
            expired = [s for s, t in _ns_last_access.items() if t < cutoff]
            for s in expired:
                _namespaces.pop(s, None)
                _ns_last_access.pop(s, None)
        if expired:
            logger.info("Namespaces expirados removidos: %s", expired)


_MAX_RESULT_ROWS = 15  # acima disso, resultado de DataFrame é descrito, não exibido completo


def _df_describe(df: pd.DataFrame, nome: str = "result") -> str:
    """Descreve um DataFrame sem expor os dados completos ao LLM."""
    dtypes = ", ".join(f"{c}: {str(t)}" for c, t in df.dtypes.items())
    sample = df.head(1).to_markdown(index=False)
    return (
        f"`{nome}` — DataFrame com {len(df)} linhas e {len(df.columns)} colunas\n"
        f"Colunas e tipos: {dtypes}\n"
        f"Primeira linha de exemplo:\n{sample}"
    )


def _ns_summary(ns: dict) -> str:
    lines = []
    for k, v in ns.items():
        if k.startswith("_") or k in _BUILTINS_SKIP:
            continue
        if isinstance(v, pd.DataFrame):
            dtypes = ", ".join(f"{c}:{str(t)}" for c, t in v.dtypes.items())
            lines.append(f"  {k}: DataFrame — {len(v)} linhas | {dtypes}")
        elif isinstance(v, pd.Series):
            lines.append(f"  {k}: Series — {len(v)} elementos, dtype: {v.dtype}")
        elif isinstance(v, (int, float, bool)):
            lines.append(f"  {k} = {v}")
        elif isinstance(v, str):
            preview = v[:60].replace("\n", " ")
            lines.append(f"  {k} = '{preview}{'…' if len(v) > 60 else ''}'")
        else:
            lines.append(f"  {k}: {type(v).__name__}")
    if not lines:
        return "\n---\n_Ambiente vazio — nenhuma variável definida ainda._"
    return "\n---\n**Variáveis disponíveis no ambiente** (criadas em passos anteriores):\n" + "\n".join(lines)


# ── Extração de thinking / texto de mensagens Gemini ─────────────────────────

def _extract_thinking(msg) -> str:
    content = getattr(msg, "content", "")
    if isinstance(content, list):
        parts = []
        for p in content:
            if not isinstance(p, dict):
                continue
            # formato Gemini thought summaries: part.thought == True
            if p.get("thought") or p.get("type") == "thinking":
                text = p.get("text") or p.get("thinking", "")
                if text:
                    parts.append(str(text))
        return "\n".join(parts)
    return getattr(msg, "additional_kwargs", {}).get("reasoning_content", "")


def _extract_text_content(msg) -> str:
    content = getattr(msg, "content", "")
    if isinstance(content, list):
        return "\n".join(
            str(p.get("text", "")) for p in content
            if isinstance(p, dict) and p.get("type") in ("text", None)
        )
    return content if isinstance(content, str) else ""


def _emit_agent_events(response, push_fn) -> None:
    """Extrai thinking, texto pré-tool e tool calls de qualquer resposta de agente
    e emite eventos via push_fn(event_dict)."""
    thinking = _extract_thinking(response)
    if thinking:
        push_fn({"type": "thinking", "text": thinking})

    tool_calls = getattr(response, "tool_calls", None) or []
    fc = getattr(response, "additional_kwargs", {}).get("function_call")
    if not tool_calls and fc:
        tool_calls = [{"name": fc.get("name", "")}]

    if tool_calls:
        pre_text = _extract_text_content(response)
        if pre_text:
            push_fn({"type": "thinking", "text": pre_text})
        for tc in tool_calls:
            name = tc.get("name", "")
            label = _TOOL_LABELS.get(name, f"⚙ {name}")
            push_fn({"type": "tool", "name": name, "label": label})


# ── Side-channel de eventos do sub-agente ─────────────────────────────────────
# O sub-agente roda dentro de um @tool (analisar_grafico), fora do stream do
# orquestrador. Usamos uma fila por sessão para capturar thinking e tool calls
# do sub-agente e entregá-los ao stream principal após cada nó completar.

_event_queues: dict = {}  # session_id → queue.Queue
_eq_lock = threading.Lock()

_TOOL_LABELS: dict[str, str] = {
    "read_skill":              "📖 Lendo instruções da skill",
    "calcular_periodo":        "📅 Calculando período",
    "chamar_api":              "🌐 Buscando dados da API",
    "executar_sql":            "🗄️ Consultando banco de dados",
    "analisar_dataframe":      "🔢 Processando e analisando dados",
    "consultar_analista":      "🔍 Delegando ao sub-agente analista",
    "rag_dominio":             "📚 Consultando base de conhecimento (domínio)",
    "rag_capacidades":         "📚 Consultando base de conhecimento (capacidades)",
    "rag_arquitetura":         "📚 Consultando base de conhecimento (arquitetura)",
    "rag_dados":               "📚 Consultando base de conhecimento (dados)",
    "get_current_datetime":    "🕐 Verificando data e hora",
    "gerar_pdf":               "📄 Gerando relatório PDF",
    "gerar_excel":             "📊 Gerando planilha Excel",
    "gerenciar_agenda":          "🗓️ Gerenciando agenda de relatórios",
    "criar_tarefa_agendada":    "🗓️ Criando tarefa agendada",
    "editar_tarefa_agendada":   "✏️ Editando tarefa agendada",
    "schedule_task":            "🗓️ Agendando tarefa",
    "set_task_instructions":    "📝 Salvando instruções da tarefa",
    "get_task_instructions":    "🔍 Lendo instruções da tarefa",
    "list_scheduled_tasks":     "📋 Listando tarefas agendadas",
    "delete_scheduled_task":    "🗑️ Removendo tarefa agendada",
    "update_scheduled_task":    "✏️ Editando tarefa agendada",
    "toggle_pause_task":        "⏸️ Pausando/retomando tarefa",
    "get_task_code":            "🔍 Lendo código da tarefa",
    "test_task_code":           "🧪 Testando código da tarefa",
    "save_task_code":           "💾 Salvando código da tarefa",
    "get_task_code_versions":   "📜 Listando versões do código",
    "restore_task_code_version": "↩️ Restaurando versão do código",
    "test_widget_code":          "🧪 Testando código do painel customizado",
    "add_chart_to_dashboard":    "📌 Adicionando painel ao dashboard",
    "list_dashboard_widgets":    "📋 Listando painéis customizados",
    "delete_dashboard_widget":   "🗑️ Removendo painel do dashboard",
    "get_widget_code":           "🔍 Lendo código do painel",
    "update_widget":             "✏️ Atualizando painel do dashboard",
}


def _push_event(session_id: str, event: dict) -> None:
    with _eq_lock:
        q = _event_queues.get(session_id)
    if q is not None:
        q.put(event)


# ── Estados compartilhados ────────────────────────────────────────────────────

class State(TypedDict):
    messages: Annotated[list[BaseMessage], add_messages]


class SchedulingState(TypedDict):
    mode: str          # "create" | "edit"
    session_id: str
    user_request: str  # original user request / modification description
    task_id: str       # empty until created; for edit, the target task
    task_name: str     # task name (for create)
    task_code: str     # code being built or corrected
    analista_context: str  # skill + endpoint + colunas + amostra de dados (create only)
    error: str         # last test error; empty on success
    retries: int       # correction attempts so far
    result: str        # final output message


# ── Geração de PDF ───────────────────────────────────────────────────────────

import re as _re
import base64 as _b64

_CHART_TOKEN_RE = _re.compile(r'\[chart:([a-f0-9\-]{36})\]')


def _safe(text: str) -> str:
    """Converte para latin-1 para compatibilidade com fontes core do fpdf2."""
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _pdf_render_table(pdf, rows: list[list[str]]) -> None:
    if not rows:
        return
    col_n = max(len(r) for r in rows)
    if col_n == 0:
        return
    # Máximo de colunas que cabem com largura mínima de 20mm
    max_cols = max(1, int(185 / 20))
    col_n = min(col_n, max_cols)
    col_w = 185 / col_n
    for i, row in enumerate(rows):
        if i == 0:
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_fill_color(20, 28, 39)
            pdf.set_text_color(226, 232, 240)
        else:
            pdf.set_font("Helvetica", "", 8)
            fill = (10, 14, 20) if i % 2 == 0 else (14, 20, 30)
            pdf.set_fill_color(*fill)
            pdf.set_text_color(148, 163, 184)
        pdf.set_x(pdf.l_margin)
        for cell in row[:col_n]:
            text = _safe(cell)
            while text and pdf.get_string_width(text) > col_w - 1:
                text = text[:-1]
            pdf.cell(col_w, 5.5, text, border=1, fill=True)
        pdf.ln()
    pdf.set_x(pdf.l_margin)
    pdf.ln(3)


def _pdf_safe_multicell(pdf, h: float, text: str) -> None:
    """Chama multi_cell garantindo que x está no l_margin e quebrando palavras longas."""
    pdf.set_x(pdf.l_margin)
    # Quebra palavras sem espaço que ultrapassem a largura disponível
    available = pdf.w - pdf.l_margin - pdf.r_margin
    words = text.split(" ")
    safe_words = []
    for w in words:
        if pdf.get_string_width(w) > available - 2:
            # Quebra o token longo em pedaços que caibam
            chunk, buf = "", ""
            for ch in w:
                if pdf.get_string_width(buf + ch) > available - 2:
                    safe_words.append(buf)
                    buf = ch
                else:
                    buf += ch
            if buf:
                safe_words.append(buf)
        else:
            safe_words.append(w)
    pdf.multi_cell(0, h, _safe(" ".join(safe_words)))


def _pdf_render_text(pdf, text: str) -> None:
    table_rows: list[list[str]] = []

    def flush_table():
        if table_rows:
            _pdf_render_table(pdf, table_rows)
            table_rows.clear()

    for line in text.split("\n"):
        s = line.strip()
        if not s:
            flush_table()
            pdf.ln(2)
            continue

        # Linha separadora de tabela (|---|---|)
        if s.startswith("|") and s.endswith("|") and _re.fullmatch(r'[\|\-\: ]+', s):
            continue

        # Linha de tabela
        if s.startswith("|") and s.endswith("|"):
            cells = [c.strip() for c in s[1:-1].split("|")]
            table_rows.append(cells)
            continue

        flush_table()

        if s.startswith("### "):
            pdf.set_font("Helvetica", "B", 11)
            pdf.set_text_color(226, 232, 240)
            _pdf_safe_multicell(pdf, 6, s[4:])
            pdf.ln(1)
        elif s.startswith("## "):
            pdf.set_font("Helvetica", "B", 13)
            pdf.set_text_color(56, 189, 248)
            _pdf_safe_multicell(pdf, 7, s[3:])
            pdf.ln(2)
        elif s.startswith("# "):
            pdf.set_font("Helvetica", "B", 15)
            pdf.set_text_color(56, 189, 248)
            _pdf_safe_multicell(pdf, 8, s[2:])
            pdf.ln(3)
        elif s.startswith(("- ", "* ")):
            clean = _re.sub(r'\*\*(.*?)\*\*', r'\1', s[2:])
            clean = _re.sub(r'`(.*?)`', r'\1', clean)
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(148, 163, 184)
            _pdf_safe_multicell(pdf, 5, f"  • {clean}")
        else:
            clean = _re.sub(r'\*\*(.*?)\*\*', r'\1', s)
            clean = _re.sub(r'\*(.*?)\*', r'\1', clean)
            clean = _re.sub(r'`(.*?)`', r'\1', clean)
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(148, 163, 184)
            _pdf_safe_multicell(pdf, 5, clean)

    flush_table()


def _build_pdf(titulo: str, conteudo: str, session_id: str) -> str:
    from fpdf import FPDF

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Cabeçalho
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(56, 189, 248)
    _pdf_safe_multicell(pdf, 10, titulo)
    pdf.ln(2)
    pdf.set_draw_color(30, 45, 61)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(5)

    # Divide conteúdo em partes de texto e tokens de chart
    try:
        parts = _CHART_TOKEN_RE.split(conteudo)
        for i, part in enumerate(parts):
            if i % 2 == 0:
                _pdf_render_text(pdf, part)
            else:
                b64 = chart_store.get_chart_b64(part)
                if b64:
                    img_bytes = _b64.b64decode(b64)
                    if pdf.get_y() > 210:
                        pdf.add_page()
                    pdf.image(io.BytesIO(img_bytes), x=10, w=190)
                    pdf.ln(3)
    except Exception:
        # Fallback: página limpa com mensagem de erro de layout
        pdf.add_page()
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(148, 163, 184)
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(0, 6, _safe("Erro ao renderizar conteudo formatado. Verifique o conteudo gerado."))

    pdf_bytes = bytes(pdf.output())
    safe_name = _re.sub(r'[^\w\- ]', '', titulo)[:40].strip().replace(" ", "_")
    filename = f"{safe_name or 'relatorio'}.pdf"
    return chart_store.save_pdf(session_id, pdf_bytes, filename)


# ── Tools do orquestrador ─────────────────────────────────────────────────────

_RAG_DIR = Path(__file__).parent / "rag"


@tool
def rag_dominio() -> str:
    """Retorna informações sobre o domínio industrial E o mapeamento ATUAL linha→modelo
    lido ao vivo do banco de dados.

    CHAME ESTA TOOL SEMPRE QUE o usuário mencionar um modelo pelo nome
    (ex: "PhoneX Pro", "PhoneX Lite", "PhoneX Ultra", "PhoneX Mini") para descobrir
    qual line= usar nos filtros da API. O mapeamento pode mudar — não assuma valores fixos.

    Também use para: definição dos KPIs (OEE, FPY, downtime, taxa de defeito),
    o que cada painel do dashboard exibe, como interpretar os indicadores,
    turnos (A/B/C), granularidade dos dados (diária vs horária).
    """
    return (_RAG_DIR / "dominio.md").read_text(encoding="utf-8")


@tool
def rag_capacidades() -> str:
    """Retorna o catálogo completo de capacidades do agente: tipos de análise disponíveis,
    exemplos de perguntas que podem ser feitas, como solicitar gráficos/tabelas/PDFs/Excel,
    como criar relatórios periódicos agendados, como configurar monitores com alertas,
    frequências suportadas e como gerenciar tarefas agendadas (listar, pausar, deletar, editar).

    Use quando o usuário perguntar: o que você pode fazer, que tipo de análise é possível,
    como agendar um relatório, como criar um alerta, quais frequências existem, como editar
    uma tarefa, como funciona o agendamento, etc.
    """
    return (_RAG_DIR / "capacidades.md").read_text(encoding="utf-8")


@tool
def rag_arquitetura() -> str:
    """Retorna a arquitetura técnica do sistema de agentes: como o orquestrador roteia
    pedidos, como o sub-agente analista executa análises (read_skill → chamar_api →
    analisar_dataframe), o catálogo de skills disponíveis, como o sub-agente de scheduling
    gerencia tarefas, os dois modos de execução (LLM vs determinístico) e os métodos do ctx.

    Use quando o usuário perguntar: como o sistema funciona internamente, quais skills
    existem, como funciona o agendamento por baixo, o que é o modo determinístico,
    o que é o sub-agente analista, como o orquestrador decide o que fazer, etc.
    """
    return (_RAG_DIR / "arquitetura.md").read_text(encoding="utf-8")


@tool
def rag_dados() -> str:
    """Retorna a referência técnica de dados: endpoints da API REST com seus filtros
    (from, to, shift, line), estrutura dos payloads e schema completo do banco SQLite
    (tabelas production, defects, metrics, hourly_production, lines_status, alerts, kpis)
    com colunas, tipos e descrições.

    Use quando o usuário perguntar: quais endpoints existem, quais filtros a API aceita,
    quais colunas tem a tabela de produção, qual o schema do banco, como cruzar tabelas,
    quais campos estão disponíveis para análise, etc.
    """
    return (_RAG_DIR / "dados.md").read_text(encoding="utf-8")

@tool
def gerar_pdf(titulo: str, conteudo: str) -> str:
    """Gera PDF. NÃO usar para Excel, planilha, .xlsx ou spreadsheet — use gerar_excel nesses casos.

    Use APENAS quando o usuário mencionar explicitamente "PDF" ou ".pdf".
    Deve ser chamado APÓS consultar_analista ter retornado a análise completa.

    Fluxo correto:
      1. Chame consultar_analista() para obter a análise com gráficos.
      2. Chame gerar_pdf() passando o resultado completo como conteudo.
      3. Responda ao usuário incluindo o token [pdf:uuid] retornado — ele será
         automaticamente convertido em link de download.

    REGRA DE CONTEÚDO — o argumento `conteudo` deve conter APENAS:
      - Títulos de seções (ex: "## Produção por Linha")
      - Tokens de gráficos gerados: [chart:uuid]
      - Tabelas de dados em markdown
      - Conclusões analíticas objetivas (números, tendências, comparações)

    PROIBIDO no conteudo — NÃO inclua NENHUMA dessas formas de texto:
      - Abertura: "Aqui está o gráfico...", "Segue abaixo...", "Conforme solicitado...", "Claro!"
      - Fechamento: "Gerado a partir dos dados...", "Relatório gerado em...", "Espero que ajude!",
        "Qualquer dúvida estou à disposição.", "Este relatório foi gerado automaticamente."
      - Confirmações/meta-comentários: "Análise concluída!", "O PDF foi gerado.", "Pronto!"
      - Qualquer frase que fale sobre o próprio relatório em vez de conter dados do relatório

    Exemplos CORRETOS de conteudo:
      "## Produção Semanal\\n[chart:abc-123]\\n\\nLinha 1 liderou com 1.240 un. (+8% vs meta).\\nFPY médio: 96,2%. Pior dia: segunda (91%)."

    Exemplos ERRADOS (nunca incluir):
      INÍCIO: "Aqui está a análise de produção que você pediu.\\n## Produção Semanal\\n..."
      FIM: "...FPY médio: 96,2%.\\n\\nGerado a partir dos dados do sistema MFG."
      FIM: "...FPY médio: 96,2%.\\n\\nEspero que o relatório seja útil!"

    Args:
        titulo: Título do relatório (ex: "Análise de Produção — Maio 2026").
        conteudo: Análise em markdown com tokens [chart:uuid], tabelas e conclusões.
                  Sem texto introdutório ou conversacional.
    """
    _tlog("gerar_pdf", "CHAMADA", titulo=titulo, conteudo_chars=len(conteudo))
    try:
        pdf_id = _build_pdf(titulo, conteudo, _current_session.get())
        result = f"[pdf:{pdf_id}]"
        _tlog("gerar_pdf", "RETORNO", status="OK", pdf_id=pdf_id)
        return result
    except Exception as e:
        msg = f"Erro ao gerar PDF: {e}"
        _tlog("gerar_pdf", "RETORNO", status="ERRO", erro=msg)
        return msg


@tool
def gerar_excel(nome_dataframe: str, nome_arquivo: str, sheets: Optional[str] = None) -> str:
    """Cria um arquivo Excel (.xlsx) e disponibiliza link para download.

    Use quando o usuário pedir Excel, planilha, xlsx ou spreadsheet.
    Não use para PDF (use gerar_pdf) nem para exibir tabela na tela.

    Fluxo correto:
      1. Chame consultar_analista(tipo='tabela') para carregar os dados no namespace.
         O sub-agente retornará o nome do DataFrame (ex: 'producao', 'resultado').
      2. Chame gerar_excel() passando o nome do DataFrame.
      3. Responda ao usuário incluindo o token [excel:uuid] retornado — ele será
         automaticamente convertido em link de download.

    Para múltiplas abas, use sheets='Aba1=df1,Aba2=df2'.

    Args:
        nome_dataframe: Nome do DataFrame no namespace (ex: 'resultado'). Ignorado se sheets fornecido.
        nome_arquivo: Nome do arquivo sem extensão (ex: 'producao_maio').
        sheets: Opcional. Múltiplas abas no formato 'Aba1=df1,Aba2=df2'.
    """
    import io
    _tlog("gerar_excel", "CHAMADA", nome_dataframe=nome_dataframe, nome_arquivo=nome_arquivo, sheets=sheets)
    try:
        ns = _ns_get()
        output = io.BytesIO()

        if sheets and sheets.strip():
            # múltiplas abas: "Aba1=df1,Aba2=df2"
            mapa = {}
            for par in sheets.split(","):
                par = par.strip()
                if "=" not in par:
                    continue
                aba, var = par.split("=", 1)
                mapa[aba.strip()] = var.strip()

            if not mapa:
                return "Erro: parâmetro sheets inválido. Use formato 'Aba1=df1,Aba2=df2'."

            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                for aba, var in mapa.items():
                    df = ns.get(var)
                    if df is None or not isinstance(df, pd.DataFrame):
                        return f"Erro: variável '{var}' não encontrada ou não é um DataFrame."
                    df.to_excel(writer, sheet_name=aba[:31], index=False)
                    _fmt_excel_sheet(writer.sheets[aba[:31]], df)
        else:
            df = ns.get(nome_dataframe)
            if df is None or not isinstance(df, pd.DataFrame):
                return (
                    f"Erro: variável '{nome_dataframe}' não encontrada no namespace. "
                    f"Variáveis disponíveis: {[k for k, v in ns.items() if isinstance(v, pd.DataFrame)]}"
                )
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Dados", index=False)
                _fmt_excel_sheet(writer.sheets["Dados"], df)

        excel_bytes = output.getvalue()
        safe_name = _re.sub(r'[^\w\- ]', '', nome_arquivo)[:40].strip().replace(" ", "_")
        filename = f"{safe_name or 'dados'}.xlsx"
        excel_id = chart_store.save_excel(_current_session.get(), excel_bytes, filename)
        result = f"[excel:{excel_id}]"
        _tlog("gerar_excel", "RETORNO", status="OK", excel_id=excel_id, filename=filename)
        return result
    except Exception as e:
        msg = f"Erro ao gerar Excel: {e}"
        _tlog("gerar_excel", "RETORNO", status="ERRO", erro=msg)
        return msg


def _fmt_excel_sheet(ws, df: pd.DataFrame) -> None:
    """Aplica largura automática e cabeçalho destacado na worksheet openpyxl."""
    from openpyxl.styles import PatternFill, Font, Alignment

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

        # auto-width: maior entre o cabeçalho e o conteúdo da coluna
        max_len = len(str(col_name))
        for row_idx in range(2, min(len(df) + 2, 1002)):  # amostra de até 1000 linhas
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 60)


@tool
def get_current_datetime() -> str:
    """Retorna a data e hora atual do sistema.

    Use esta tool sempre que o usuário perguntar que horas são,
    qual é a data de hoje, ou qualquer variação dessas perguntas.
    """
    result = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    _tlog("get_current_datetime", "RETORNO", resultado=result)
    return result


@tool
def get_dashboard_charts() -> str:
    """Retorna os dados dos gráficos e KPIs atualmente exibidos no dashboard do usuário.

    Use esta tool quando o usuário perguntar sobre o que está vendo na tela,
    mencionar valores, tendências ou comparações visíveis nos gráficos,
    ou usar expressões como 'no dashboard', 'na tela', 'o que está aparecendo',
    'o gráfico mostra', 'os KPIs', 'os alertas ativos'.
    NÃO use para análises que exijam buscar dados históricos — use consultar_analista nesses casos.
    """
    session = _current_session.get()
    with _cs_lock:
        snapshot = _chart_snapshots.get(session)

    if not snapshot:
        return "Nenhum snapshot do dashboard disponível para esta sessão."

    lines: list[str] = []

    kpis = snapshot.get("kpis", [])
    if kpis:
        lines.append("=== KPIs ===")
        for k in kpis:
            badge = f" [{k.get('badge', '')}]" if k.get("badge") else ""
            lines.append(f"  {k['label']}: {k['value']}{badge}")
        lines.append("")

    charts = snapshot.get("charts", {})
    for chart_id, chart in charts.items():
        title = chart.get("title", chart_id)
        labels = chart.get("labels", [])
        datasets = chart.get("datasets", [])
        lines.append(f"=== {title} ===")
        if labels:
            lines.append(f"  Labels: {', '.join(str(l) for l in labels)}")
        for ds in datasets:
            ds_label = ds.get("label", "")
            data_str = ", ".join(str(v) for v in ds.get("data", []))
            lines.append(f"  {ds_label}: {data_str}")
        lines.append("")

    alerts = snapshot.get("alerts", [])
    if alerts:
        lines.append(f"=== Alertas ({len(alerts)} total) ===")
        for a in alerts[:20]:
            val = a.get("value")
            thr = a.get("threshold")
            ts = a.get("created_at", "")
            detail = f" (valor: {val}, limiar: {thr})" if val is not None else ""
            read_flag = " [lido]" if a.get("read") else " [novo]"
            lines.append(f"  [{ts}]{read_flag} {a.get('message', '')}{detail}")
    else:
        lines.append("=== Alertas ===\n  Nenhum alerta ativo na tela.")
    lines.append("")

    artifacts = snapshot.get("artifacts", [])
    if artifacts:
        lines.append(f"=== Artifacts ({len(artifacts)} total) ===")
        for a in artifacts:
            lines.append(f"  [{a.get('type', '?')}] {a.get('label', a.get('id', ''))} — {a.get('created_at', '')}")
    else:
        lines.append("=== Artifacts ===\n  Nenhum artifact gerado.")
    lines.append("")

    tasks = snapshot.get("tasks", [])
    if tasks:
        lines.append(f"=== Tarefas Agendadas ({len(tasks)} total) ===")
        for t in tasks:
            lines.append(f"  [#{t.get('id', '?')}] {t.get('name', '')} — status: {t.get('status', '')} — schedule: {t.get('schedule', '')}")
    else:
        lines.append("=== Tarefas Agendadas ===\n  Nenhuma tarefa agendada.")

    _tlog("get_dashboard_charts", "RETORNO", linhas=len(lines))
    return "\n".join(lines) if lines else "Dashboard sem dados disponíveis."


# ── Tools do sub-agente ───────────────────────────────────────────────────────

@tool
def read_skill(filename: str) -> str:
    """Lê as instruções completas de uma skill antes de executar qualquer análise.

    OBRIGATÓRIO: chame esta tool SEMPRE como primeiro passo de qualquer análise,
    antes de chamar_api, executar_sql ou analisar_dataframe. A skill define quais
    endpoints usar, quais parâmetros passar e como processar os dados.

    Args:
        filename: Nome do arquivo .md da skill (ex: 'analise_producao.md').
                  Consulte o catálogo no system prompt para ver os arquivos disponíveis.
    """
    _tlog("read_skill", "CHAMADA", filename=filename)
    caminho = SKILLS_FOLDER / filename
    if not caminho.exists():
        arquivos = [f for f in os.listdir(SKILLS_FOLDER) if f.endswith(".md")]
        msg = f"Arquivo '{filename}' não encontrado. Skills disponíveis: {arquivos}"
        _tlog("read_skill", "RETORNO", resultado=msg)
        return msg
    conteudo = caminho.read_text(encoding="utf-8")
    _tlog("read_skill", "RETORNO", chars=len(conteudo), conteudo=conteudo)
    return conteudo


@tool
def chamar_api(url: str, chave: str, params: Optional[dict] = None) -> str:
    """Busca dados de uma API REST e injeta o DataFrame no ambiente de análise.

    O DataFrame fica disponível como variável com o nome da chave (ex: chave='producao'
    → variável `producao` no ambiente). Nunca retorna os dados ao LLM.

    Args:
        url: URL completa do endpoint (ex: 'http://localhost:8000/production').
        chave: Nome da variável no ambiente (ex: 'producao', 'defeitos').
        params: Parâmetros de query opcionais (ex: {"from": "2026-05-01", "to": "2026-05-15"}).
    """
    _tlog("chamar_api", "CHAMADA", url=url, chave=chave, params=params)
    try:
        response = http_requests.get(url, params=params, timeout=15)
        response.raise_for_status()
        data = response.json()
        df = pd.DataFrame(data) if isinstance(data, list) else pd.DataFrame([data])
        ns = _ns_get()
        ns[chave] = df
        msg = f"DataFrame '{chave}' injetado no ambiente: {len(df)} linhas, colunas: {list(df.columns)}."
        _tlog("chamar_api", "RETORNO", status="OK", linhas=len(df), colunas=list(df.columns))
    except http_requests.exceptions.HTTPError as e:
        msg = f"Erro HTTP {e.response.status_code}: {e.response.text}"
        _tlog("chamar_api", "RETORNO", status="ERRO", erro=msg)
        return msg
    except Exception as e:
        msg = f"Erro ao chamar API '{url}': {e}"
        _tlog("chamar_api", "RETORNO", status="ERRO", erro=msg)
        return msg
    return msg + _ns_summary(ns)


@tool
def executar_sql(query: str, chave: str) -> str:
    """Executa uma query SQL SELECT no banco PostgreSQL (schema brazil) e injeta o resultado como DataFrame no ambiente.

    Use esta tool quando nenhuma outra skill cobrir o pedido — análises ad-hoc que precisam
    cruzar tabelas (JOIN), rankings, ou qualquer consulta que a API não oferece.

    O search_path é fixado em 'brazil', então você pode referenciar as tabelas sem prefixo:
      purchase_order, order_item, product, customer, city
    Ou com prefixo explícito: brazil.purchase_order, etc.

    O DataFrame fica disponível como variável com o nome da chave (ex: chave='resultado'
    → variável `resultado` no ambiente para uso em analisar_dataframe).

    IMPORTANTE: apenas SELECT é permitido. Qualquer outra instrução será rejeitada.
    Não use SELECT * — liste apenas as colunas necessárias.

    Args:
        query: Query SQL SELECT a ser executada (ex: 'SELECT status::text, COUNT(*) FROM purchase_order GROUP BY status').
        chave: Nome da variável no ambiente onde o DataFrame será armazenado (ex: 'resultado', 'dados').
    """
    _tlog("executar_sql", "CHAMADA", query=query, chave=chave)

    normalized = query.strip().lstrip("(").upper()
    if not normalized.startswith("SELECT"):
        msg = "Apenas queries SELECT são permitidas. Instrução rejeitada."
        _tlog("executar_sql", "RETORNO", status="ERRO", erro=msg)
        return msg

    try:
        import psycopg2
        conn = psycopg2.connect(
            host=os.getenv("POSH_DB_HOST", "127.0.0.1"),
            port=int(os.getenv("POSH_DB_PORT", "5432")),
            user=os.getenv("POSH_DB_USER", "postgres"),
            password=os.getenv("POSH_DB_PASSWORD", "Moto#1234"),
            dbname=os.getenv("POSH_DB_NAME", "postgres"),
            options="-c search_path=brazil -c default_transaction_read_only=on",
        )
        try:
            df = pd.read_sql_query(query, conn)
        finally:
            conn.close()
        ns = _ns_get()
        ns[chave] = df
        msg = (
            f"DataFrame '{chave}' injetado no ambiente: {len(df)} linhas, "
            f"colunas: {list(df.columns)}."
        )
        _tlog("executar_sql", "RETORNO", status="OK", linhas=len(df), colunas=list(df.columns))
    except Exception as e:
        msg = f"Erro ao executar SQL: {e}"
        _tlog("executar_sql", "RETORNO", status="ERRO", erro=msg)
        return msg

    return msg + _ns_summary(ns)


@tool
def analisar_dataframe(script: str) -> str:
    """Processa e analisa os dados disponíveis no ambiente, retornando o resultado formatado.

    Variáveis carregadas por chamar_api estão disponíveis pelo nome da chave.
    Variáveis criadas em chamadas anteriores desta tool também estão disponíveis.
    Atribua à variável `result` o que deve ser exibido — DataFrames são automaticamente
    convertidos para tabela markdown. Use print() para textos adicionais.

    IMPORTANTE — nomeação de variáveis para geração de Excel:
    Quando o resultado será usado pelo orquestrador para gerar_excel, atribua o DataFrame
    a uma variável com nome descritivo no script (ex: `producao = df_filtrado`).
    Após executar, mencione explicitamente na resposta final o nome da variável
    (ex: "DataFrame 'producao' disponível no namespace com 30 linhas.").

    Args:
        script: Lógica de análise a executar sobre os dados disponíveis.
    """
    _tlog("analisar_dataframe", "CHAMADA", script=script)

    ns = _ns_get()

    if _sandbox_mode.get():
        import re as _re_sb
        # Remove imports explícitos — tudo já está pré-injetado no namespace do sandbox.
        script = "\n".join(
            ln for ln in script.splitlines()
            if not _re_sb.match(r'\s*(import |from \S+ import )', ln)
        )
        from scheduler.runner import _build_globals
        from scheduler.context import TaskContext
        globals_exec = _build_globals(
            TaskContext(_current_session.get()), "2000-01-01", "2000-01-01"
        )
        globals_exec.update({k: v for k, v in ns.items() if isinstance(v, pd.DataFrame)})
    else:
        globals_exec = ns

    old_stdout = sys.stdout
    sys.stdout = io.StringIO()
    try:
        exec(textwrap.dedent(script).lstrip("\n"), globals_exec)
        console = sys.stdout.getvalue().strip()
        result = globals_exec.get("result")

        parts = []
        if console:
            parts.append(f"```\n{console}\n```")
        if result is not None:
            if isinstance(result, matplotlib.figure.Figure):
                buf = io.BytesIO()
                result.savefig(
                    buf, format="png", dpi=130, bbox_inches="tight",
                    facecolor="white", edgecolor="none",
                )
                buf.seek(0)
                chart_id = chart_store.save_chart(_current_session.get(), buf.getvalue())
                plt.close(result)
                parts.append(f"[chart:{chart_id}]")
            elif isinstance(result, pd.DataFrame):
                if len(result) <= _MAX_RESULT_ROWS:
                    parts.append(result.to_markdown(index=False))
                else:
                    parts.append(_df_describe(result))
            elif isinstance(result, pd.Series):
                if len(result) <= _MAX_RESULT_ROWS:
                    parts.append(result.to_frame().to_markdown())
                else:
                    parts.append(_df_describe(result.to_frame()))
            else:
                parts.append(str(result))
        if not parts:
            parts.append("Script executado (sem saída em `result`).")

        history = ns.get("_script_history", [])
        history.append({"query": ns.get("_current_query", ""), "script": script})
        ns["_script_history"] = history[-3:]

        parts.append(_ns_summary(ns))
        saida = "\n\n".join(parts)
        _tlog("analisar_dataframe", "RETORNO", status="OK", saida=saida)
        return saida

    except Exception as e:
        msg = f"Erro:\n{type(e).__name__}: {e}\n\n{tb.format_exc()}"
        msg += _ns_summary(ns)
        _tlog("analisar_dataframe", "RETORNO", status="ERRO", erro=str(e))
        if _sandbox_mode.get():
            try:
                from db import log_task_code_error
                log_task_code_error(
                    task_id=_sandbox_task_id.get() or "unknown",
                    attempt=0,
                    phase="explore",
                    error=f"{type(e).__name__}: {e}",
                    code=script,
                )
            except Exception:
                pass
        return msg
    finally:
        sys.stdout = old_stdout


@tool
def calcular_periodo(periodo: str) -> str:
    """Converte um período em linguagem natural para from_date e to_date em YYYY-MM-DD.

    OBRIGATÓRIO: chame esta tool ANTES de consultar_analista. O resultado fornece
    os parâmetros from_date e to_date exigidos por consultar_analista — sem eles
    a consulta ao sub-agente analista não pode ser realizada.

    Retorna uma string no formato: from=YYYY-MM-DD&to=YYYY-MM-DD

    Args:
        periodo: Descrição do período desejado. Exemplos aceitos:
            'hoje', 'ontem', 'esta semana', 'semana passada',
            'ultimos 7 dias', 'ultimos 14 dias', 'ultimos 30 dias',
            'este mes', 'mes passado'.
    """
    _tlog("calcular_periodo", "CHAMADA", periodo=periodo)
    hoje = datetime.now().date()
    p = periodo.lower().strip()

    if p in ("hoje", "today"):
        frm = to = hoje
    elif p in ("ontem", "yesterday"):
        frm = to = hoje - timedelta(days=1)
    elif p in ("esta semana", "essa semana", "this week"):
        frm = hoje - timedelta(days=hoje.weekday())
        to = hoje
    elif p in ("semana passada", "last week"):
        start = hoje - timedelta(days=hoje.weekday() + 7)
        frm = start
        to = start + timedelta(days=6)
    elif p in ("este mes", "esse mes", "este mês", "esse mês", "this month"):
        frm = hoje.replace(day=1)
        to = hoje
    elif p in ("mes passado", "mês passado", "last month"):
        primeiro_deste = hoje.replace(day=1)
        to = primeiro_deste - timedelta(days=1)
        frm = to.replace(day=1)
    else:
        # Extrai número de dias de expressões como "ultimos 7 dias", "last 30 days"
        import re
        m = re.search(r"(\d+)", p)
        if m:
            days = int(m.group(1))
            frm = hoje - timedelta(days=days - 1)
            to = hoje
        else:
            msg = (
                f"Período '{periodo}' não reconhecido. "
                "Use: 'hoje', 'ontem', 'esta semana', 'semana passada', "
                "'ultimos N dias', 'este mes', 'mes passado'."
            )
            _tlog("calcular_periodo", "RETORNO", status="ERRO", msg=msg)
            return msg

    result = f"from={frm.isoformat()}&to={to.isoformat()}"
    _tlog("calcular_periodo", "RETORNO", status="OK", result=result)
    return result


# ── Catálogo de skills (injetado no system prompt do sub-agente) ──────────────

def _montar_catalogo_skills() -> str:
    """Lê os headers dos arquivos .md e monta o texto de catálogo para o system prompt."""
    try:
        arquivos = sorted(f for f in os.listdir(SKILLS_FOLDER) if f.endswith(".md"))
    except FileNotFoundError:
        return "Nenhuma skill disponível."
    if not arquivos:
        return "Nenhuma skill disponível."

    linhas = ["Skills disponíveis (use read_skill para ver as instruções completas):\n"]
    for arq in arquivos:
        caminho = SKILLS_FOLDER / arq
        with open(caminho, encoding="utf-8") as f:
            header = "".join(f.readline() for _ in range(SKILL_HEADER_LINES))
        linhas.append(f"  Arquivo: {arq}\n{header.strip()}\n")

    return "\n".join(linhas)


# ── Sub-agente analista ───────────────────────────────────────────────────────

def _build_sub_agent(llm):
    catalogo = _montar_catalogo_skills()

    hoje = datetime.now().strftime("%d/%m/%Y")

    SUB_SYSTEM_PROMPT = (
        f"Data de hoje: {hoje}\n\n"
        "Você é um sub-agente especializado em análise de dados industriais.\n\n"
        "## Fluxo obrigatório\n"
        "Para QUALQUER pedido de análise, siga estes passos nesta ordem:\n"
        "  1. Chame read_skill() para obter as instruções da skill correta.\n"
        "  2a. Se a skill usa chamar_api: chame calcular_periodo() e depois chamar_api().\n"
        "  2b. Se a skill usa executar_sql (skill analise_sql_livre): chame executar_sql() diretamente "
        "com a query SQL adequada — NÃO chame calcular_periodo() nem chamar_api() neste caso.\n"
        "  3. Chame analisar_dataframe() para processar os dados e obter o resultado.\n"
        "     - Você pode chamar analisar_dataframe() várias vezes para análises em etapas.\n"
        "     - Variáveis criadas em chamadas anteriores de analisar_dataframe continuam disponíveis.\n"
        "  4. Só então redija a resposta final.\n\n"
        "## Modificação de gráfico ou análise já feita\n"
        "Se o usuário pedir para alterar algo em um gráfico ou análise anterior "
        "(cores, tipo de gráfico, estilo, filtro, agrupamento, etc.), "
        "repita o fluxo completo (passos 1-3) usando os MESMOS parâmetros da análise anterior "
        "— você os encontra no histórico da conversa — e aplique a modificação solicitada no script "
        "passado para analisar_dataframe. NUNCA responda que não é possível modificar um gráfico.\n\n"
        "## Quando usar analise_sql_livre\n"
        "Se o pedido do usuário não puder ser atendido por nenhuma skill existente "
        "(cruzamentos entre tabelas, rankings, análises customizadas), use read_skill('analise_sql_livre.md') "
        "e depois executar_sql() para construir a query SQL adequada.\n\n"
        "## Tipos de dados — invariantes globais\n"
        "- A coluna `date` em TODOS os DataFrames é sempre string (TEXT/object). "
        "Antes de qualquer operação com datas (strftime, sort por data, comparação, resample, etc.), "
        "converta obrigatoriamente: `df['date'] = pd.to_datetime(df['date'])`.\n\n"
        "## Regras\n"
        "- NUNCA invente dados — use apenas resultados de analisar_dataframe.\n"
        "- Se analisar_dataframe retornar erro, corrija e chame novamente.\n"
        "- Formate SEMPRE a resposta final em markdown: tabelas para dados tabulares, "
        "**negrito** para valores relevantes.\n"
        "- LEGENDA DOS GRÁFICOS: NUNCA use `bbox_to_anchor` para posicionar a legenda fora "
        "dos eixos — isso gera figuras extremamente largas com a legenda separada do gráfico. "
        "Use sempre `ax.legend(loc='best')` ou outra posição interna (upper right, lower left, etc). "
        "Se o gráfico tiver muitas séries, use `ax.legend(loc='upper right', fontsize=8)` "
        "dentro dos eixos. PROIBIDO: `bbox_to_anchor`, `loc='outside'`, `fig.legend()`.\n"
        "- RÓTULOS DO EIXO X: para rotacionar rótulos use SOMENTE "
        "`plt.setp(ax.get_xticklabels(), rotation=45, ha='right')`. "
        "NUNCA passe `ha=` dentro de `ax.tick_params()` — esse parâmetro não existe lá e causa TypeError.\n"
        "- TIPO DE SAÍDA: a mensagem pode começar com um prefixo `[TIPO DE SAÍDA OBRIGATÓRIO: X]`.\n"
        "  - GRAFICO → `result = fig` (figura matplotlib). NUNCA retorne tabela quando tipo=GRAFICO.\n"
        "  - TABELA  → `result = df`. Mencione na resposta final o nome da variável DataFrame\n"
        "    ex: 'DataFrame `producao` disponível no namespace com 30 linhas.'\n"
        "  - AMBOS   → gráfico em uma chamada de analisar_dataframe, tabela em outra.\n"
        "  Se não houver prefixo, padrão é GRAFICO.\n"
        "- PRINT ANTES DO GRÁFICO: no script que gera um gráfico, sempre faça print() dos "
        "valores agregados principais ANTES de atribuir `result = fig`. Exemplo:\n"
        "    print(df_grouped.to_string())  # ou print(f'Total: {total}')\n"
        "    result = fig\n"
        "  Isso garante que os valores reais apareçam no retorno de analisar_dataframe e "
        "possam ser usados no bloco **Dados do gráfico:**.\n"
        "- TOKENS DE GRÁFICO: copie LITERALMENTE o token `[chart:uuid]` retornado por "
        "analisar_dataframe na resposta final. Nunca omita — sem ele o gráfico não aparece.\n"
        "- RESUMO DO GRÁFICO: sempre que incluir um token `[chart:uuid]`, adicione logo abaixo "
        "um bloco **Dados do gráfico:** com 5 a 10 pontos-chave extraídos dos dados reais "
        "(totais, máximos, mínimos, médias, desvios relevantes, top-N). "
        "Exemplo:\n"
        "  **Dados do gráfico:**\n"
        "  - Total do período: 1.247 peças\n"
        "  - Linha com maior produção: Linha B (420 pç)\n"
        "  - Pico: 15/05 (78 pç) | Menor: 03/05 (12 pç)\n"
        "  Isso permite ao orquestrador responder follow-ups sem reanalisar os dados.\n"
        "- RESUMO DO RELATÓRIO PDF: sempre que chamar gerar_pdf(), inclua na resposta final "
        "um bloco **Dados do relatório:** com as seções geradas e os valores principais "
        "(você já tem esses dados — usou-os para compor o conteudo do PDF). Exemplo:\n"
        "  **Dados do relatório:**\n"
        "  - Seções: Produção por linha, FPY semanal, Top defeitos\n"
        "  - Período: 01/05 a 20/05/2026\n"
        "  - Total produzido: 1.247 pç | FPY médio: 96,2% | Principal defeito: Scratch (38%)\n"
        "  Isso permite ao orquestrador responder follow-ups sobre o PDF sem reabri-lo.\n"
        "- RESUMO DA PLANILHA EXCEL: sempre que chamar gerar_excel(), inclua na resposta final "
        "um bloco **Dados da planilha:** com o shape, as colunas e os valores agregados principais "
        "(você viu esses dados no retorno de analisar_dataframe). Exemplo:\n"
        "  **Dados da planilha:**\n"
        "  - Abas: Produção (180 linhas × 6 colunas), FPY (180 × 4)\n"
        "  - Colunas principais: data, linha, turno, producao, meta, fpy\n"
        "  - Total produzido: 1.247 pç | Média diária: 41,6 pç | FPY médio: 96,2%\n"
        "  Isso permite ao orquestrador responder follow-ups sobre a planilha sem reprocessar os dados.\n\n"
        "## Modo agendamento [PARA_AGENDAMENTO]\n"
        "Se a mensagem contiver [PARA_AGENDAMENTO], após a análise normal escreva um bloco\n"
        "```task_code com o código Python completo pronto para agendamento:\n\n"
        "```task_code\n"
        "def run(from_date, to_date, ctx):\n"
        "    # implementação baseada EXATAMENTE no que você executou acima\n"
        "    # endpoint real, colunas reais, lógica real — sem inventar nada\n"
        "```\n\n"
        "REGRAS CRÍTICAS do bloco task_code:\n"
        "- Baseie-se EXCLUSIVAMENTE no endpoint real, colunas reais e lógica que você observou\n"
        "- NUNCA escreva datas específicas como '2026-05-18' — use from_date/to_date ou ctx.today()\n"
        "- NUNCA inclua linhas de import — pd, np, plt, ctx etc já estão no namespace do sandbox\n"
        "- Retorne via ctx.save_chart(fig), ctx.generate_pdf(titulo, conteudo) ou ctx.generate_excel(df, nome)\n"
        "- NUNCA retorne a figura diretamente — sempre passe por ctx.save_chart(fig)\n"
        "- Para relatórios periódicos (daily/weekly/monthly): use from_date e to_date recebidos como params\n"
        "  exemplo: ctx.api(f'/brazil/purchase-orders?from={from_date}&to={to_date}')\n"
        "- Para monitores (every_Xm/every_Xh): use ctx.today() para a data\n"
        "  exemplo: hoje = ctx.today(); ctx.api(f'/brazil/orders/summary?from={hoje}&to={hoje}')\n"
        "- Se a mensagem contiver um bloco de REGRAS DO SANDBOX, siga-as com prioridade máxima\n"
        "- NÃO use df.to_markdown() no conteúdo do PDF — use apenas bullet points e tokens [chart:uuid]\n\n"
        f"## Skills disponíveis\n{catalogo}\n\n"
        "## RACIOCÍNIO OBRIGATÓRIO\n"
        "SEMPRE que for chamar uma tool, inclua na mesma resposta um texto curto explicando "
        "o que vai fazer e por quê, ANTES do bloco de função. "
        "NUNCA emita uma tool call sem texto explicativo."
    )

    sub_tools = [read_skill, chamar_api, executar_sql, analisar_dataframe, gerar_pdf, gerar_excel]
    llm_sub = llm.bind_tools(sub_tools)
    no_sub_tools = ToolNode(sub_tools)

    def no_sub_agente(state: State) -> dict:
        from langchain_core.messages import AIMessage as _AI
        ns = _ns_get()
        history = ns.get("_script_history", [])
        if history:
            lines = [
                "\n\n## Scripts anteriores desta sessão",
                "Os scripts abaixo foram executados em análises anteriores desta conversa.",
                "Se o pedido atual for uma modificação de um desses gráficos/análises, use o "
                "script correspondente como base — chame apenas analisar_dataframe() com o "
                "script modificado, sem refazer read_skill nem chamar_api.\n",
            ]
            for i, entry in enumerate(history, 1):
                lines.append(f"### Script {i} — pedido: \"{entry['query']}\"")
                lines.append(f"```python\n{entry['script']}\n```")
            extra = "\n".join(lines)
        else:
            extra = ""
        msgs = [SystemMessage(content=SUB_SYSTEM_PROMPT + extra)] + list(state["messages"])
        _tlog("sub_agente", "LLM INVOCADO", mensagens=len(msgs))
        response = llm_sub.invoke(msgs)

        meta = getattr(response, "response_metadata", {})
        finish = meta.get("finish_reason", "")

        # Empurra thinking e tool calls do sub-agente para o side-channel
        session = _current_session.get()
        _emit_agent_events(response, lambda ev: _push_event(session, ev))

        if finish == "MALFORMED_FUNCTION_CALL":
            bad = meta.get("finish_message", "")
            _tlog("sub_agente", "MALFORMED_FUNCTION_CALL detectado", tentativa=bad)
            response = _AI(content=(
                "Erro interno: o modelo tentou computar datas em código Python em vez de "
                "chamar a tool `calcular_periodo`. Use calcular_periodo() para obter as "
                f"datas before/after e depois chame chamar_api normalmente.\n\n"
                f"Tentativa inválida capturada:\n```\n{bad}\n```"
            ))

        if os.getenv("DEBUG_TOOLS", "0").strip() == "1":
            if hasattr(response, "tool_calls") and response.tool_calls:
                nomes = [tc["name"] for tc in response.tool_calls]
                logger.info(f"\n[SUB-AGENTE] → tool calls: {nomes}")
            else:
                logger.info(
                    f"\n[SUB-AGENTE] → SEM tool call"
                    f"\n  type       : {type(response).__name__}"
                    f"\n  content    : {repr(response.content)}"
                    f"\n  tool_calls : {getattr(response, 'tool_calls', 'N/A')}"
                    f"\n  add_kwargs : {getattr(response, 'additional_kwargs', {})}"
                    f"\n  response_metadata: {meta}"
                )
        return {"messages": [response]}

    def sub_para_onde(state: State) -> str:
        ultima = state["messages"][-1]
        if hasattr(ultima, "tool_calls") and ultima.tool_calls:
            return "tools"
        return END

    builder = StateGraph(State)
    builder.add_node("sub_agente", no_sub_agente)
    builder.add_node("tools", no_sub_tools)
    builder.add_edge(START, "sub_agente")
    builder.add_conditional_edges("sub_agente", sub_para_onde, {"tools": "tools", END: END})
    builder.add_edge("tools", "sub_agente")

    return builder.compile()


# ── Tool do orquestrador que dispara o sub-agente ────────────────────────────

@tool
def consultar_analista(detalhes: str, from_date: str, to_date: str, tipo: str = "grafico", para_agendamento: bool = False) -> str:
    """Delega a análise ao sub-agente analista de dados.

    O sub-agente irá identificar a skill correta, buscar os dados na API
    e retornar uma análise baseada nos dados reais.

    IMPORTANTE: Sempre chame calcular_periodo() antes para obter from_date e to_date.

    Args:
        detalhes: O que o usuário quer analisar (ex: "produção diária",
                  "produção vs meta", "defeitos por linha").
        from_date: Data inicial no formato YYYY-MM-DD. Obter via calcular_periodo().
        to_date: Data final no formato YYYY-MM-DD. Obter via calcular_periodo().
        tipo: Tipo de saída desejado. Use EXATAMENTE um dos valores abaixo:
              - "grafico"  → o sub-agente DEVE gerar um gráfico matplotlib (result = fig).
                             Use quando o usuário usou palavras como "gráfico", "chart",
                             "plot", "visualize", "visualização", "mostre", "desenhe".
              - "tabela"   → o sub-agente retorna os dados como tabela exibida na tela.
                             Use quando o usuário pediu "tabela", "lista", "dados", "números"
                             ou qualquer exibição tabular na resposta.
                             Também use como pré-passo quando gerar_excel for chamado na sequência.
              - "ambos"    → gráfico E tabela juntos. Use apenas quando o usuário pediu
                             explicitamente visualização gráfica E dados numéricos ao mesmo tempo.
                             NÃO use "ambos" para pedidos de PDF ou Excel.
              Padrão: "grafico". Em caso de dúvida, prefira "grafico".
        para_agendamento: Se True, o sub-agente inclui na resposta um bloco técnico
              ```task_context``` com endpoint real, colunas reais e código pandas executado.
              Use SEMPRE que a análise for para criar ou editar um task_code agendado.
    """
    if _sub_agent_graph is None:
        return "Sub-agente analista não inicializado."
    ns = _ns_get()
    ns["_current_query"] = detalhes
    flags = f"[TIPO DE SAÍDA OBRIGATÓRIO: {tipo.upper()}]\n[PERÍODO: from={from_date} to={to_date}]"

    # Injeta o último script como referência para edições — sem dados, só código
    last_scripts = ns.get("_script_history", [])
    if last_scripts:
        last_script = last_scripts[-1]["script"]
        flags += (
            "\n\n[ÚLTIMO SCRIPT EXECUTADO — use como base se for ajuste sobre análise anterior; "
            "ainda assim re-busque os dados normalmente antes de executar]\n"
            f"```python\n{last_script}\n```"
        )

    if para_agendamento:
        sandbox_skill = (SKILLS_FOLDER / "task_code_sandbox.md").read_text(encoding="utf-8")
        flags += (
            "\n[PARA_AGENDAMENTO]\n\n"
            "--- REGRAS DO SANDBOX (leia antes de escrever o bloco task_code) ---\n"
            + sandbox_skill
            + "\n--- FIM DAS REGRAS DO SANDBOX ---"
        )
    msg_content = f"{flags}\n{detalhes}"
    resultado = _sub_agent_graph.invoke(
        {"messages": [HumanMessage(content=msg_content)]},
        config={"configurable": {"thread_id": _current_session.get()}, "recursion_limit": 30},
    )
    content = resultado["messages"][-1].content
    if not content or not content.strip():
        return (
            "O sub-agente analista não retornou nenhuma resposta. "
            "Isso geralmente indica uma falha interna ao processar a análise. "
            "Informe o usuário e sugira reformular a pergunta."
        )
    return content


@tool
def ver_grafico(chart_id: str) -> list:
    """Visualiza um gráfico já gerado para responder perguntas de follow-up visuais.

    Use quando o usuário perguntar algo sobre um gráfico exibido anteriormente e
    os dados textuais do histórico não forem suficientes para responder
    (ex: 'por que há um pico?', 'onde está a queda?', 'o que mostra essa parte?').

    Args:
        chart_id: UUID do gráfico, extraído do token [chart:UUID] no histórico.
    """
    _tlog("ver_grafico", "CHAMADA", chart_id=chart_id)
    png_b64 = chart_store.get_chart_b64(chart_id)
    if not png_b64:
        _tlog("ver_grafico", "RETORNO", status="NOT_FOUND")
        return [{"type": "text", "text": f"Gráfico {chart_id} não encontrado."}]
    _tlog("ver_grafico", "RETORNO", status="OK", bytes=len(png_b64))
    return [
        {"type": "text", "text": "Imagem do gráfico para análise:"},
        {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{png_b64}"}},
    ]


# ── Sub-agente de scheduling ─────────────────────────────────────────────────

def _discover_scheduler_tools() -> list:
    """Retorna todas as LangChain tools definidas em scheduler.tools, em ordem de definição."""
    import inspect
    from langchain_core.tools import BaseTool
    import scheduler.tools as _mod

    tools = []
    for name, obj in inspect.getmembers(_mod):
        if isinstance(obj, BaseTool):
            tools.append(obj)
    return tools


def _describe_scheduler_tools(tools: list) -> str:
    """Gera a seção '## Suas tools' dinamicamente a partir das tools descobertas."""
    lines = []
    for t in tools:
        first_line = (t.description or "").split("\n")[0].strip()
        lines.append(f"- {t.name:<30} → {first_line}")
    return "\n".join(lines)


def _describe_task_context() -> str:
    """Gera documentação do TaskContext dinamicamente a partir das docstrings."""
    from scheduler.context import TaskContext
    import inspect

    skip = {"__init__", "tokens"}
    lines = ["Métodos disponíveis em `ctx` (objeto TaskContext injetado no run()):\n"]

    for name, method in inspect.getmembers(TaskContext, predicate=inspect.isfunction):
        if name.startswith("_") or name in skip:
            continue
        sig = inspect.signature(method)
        params = [
            str(p) for pname, p in sig.parameters.items()
            if pname not in ("self",)
        ]
        doc = (inspect.getdoc(method) or "").split("\n")[0]  # primeira linha da docstring
        lines.append(f"  ctx.{name}({', '.join(params)})")
        lines.append(f"    → {doc}")

    return "\n".join(lines)


def _build_scheduler_agent(llm):
    sch_tools = _discover_scheduler_tools()

    _sandbox_skill = (SKILLS_FOLDER / "task_code_sandbox.md").read_text(encoding="utf-8")

    SCH_SYSTEM_PROMPT = (
        "Você é o sub-agente responsável por gerenciar o agendamento de tarefas.\n\n"
        "## Suas tools\n"
        + _describe_scheduler_tools(sch_tools) + "\n\n"
        + "## Sandbox do task_code — leia antes de escrever qualquer código\n"
        + _sandbox_skill + "\n\n"
        "## Modos de execução da tarefa\n"
        "Cada tarefa pode executar em dois modos:\n"
        "  A) **Modo LLM** (sem task_code): o daemon monta um prompt com as `instructions` e chama o agente.\n"
        "  B) **Modo determinístico** (com task_code): o daemon executa o código Python diretamente,\n"
        "     sem LLM. Mais rápido, previsível e sem custo de inferência.\n"
        "O modo B é ativado automaticamente quando task_code estiver definido. Para converter uma\n"
        "tarefa para modo B, use test_task_code para validar o código e save_task_code para salvar.\n\n"
        "## Campos importantes\n"
        "- description: texto curto e legível que descreve o que a tarefa faz. "
        "Aparece na listagem para o usuário. Ex: 'Relatório semanal de OEE toda segunda às 8h.'\n"
        "- instructions: passo a passo detalhado com o código Python validado. "
        "Não aparece na listagem. Usado no modo LLM.\n"
        "- task_code: código Python com `def run(from_date, to_date, ctx)`. "
        "Quando presente, execução é determinística. Versões anteriores ficam disponíveis para rollback.\n\n"
        "## O objeto ctx — use SEMPRE estas funções, não reimplemente\n"
        "O código recebe `ctx` com os seguintes métodos. Use-os diretamente:\n\n"
        + _describe_task_context() + "\n\n"
        "  Todos os métodos de geração/salvamento retornam tokens '[tipo:uuid]' e os\n"
        "  acumulam internamente. A função run() pode retorná-los ou deixar o ctx acumulá-los.\n\n"
        "  BIBLIOTECAS DISPONÍVEIS NO NAMESPACE (NÃO use import dentro do run()):\n"
        "  pd, np, plt, mticker, date, datetime, timedelta,\n"
        "  openpyxl, Font, PatternFill, Alignment, Border, Side, get_column_letter\n"
        "  O sandbox bloqueia __import__ — qualquer `import` dentro do run() causa erro.\n\n"
        "## Salvar task_code recebido do orquestrador\n"
        "Quando receber instrução 'save_task_code task [ID]: [código]':\n"
        "1. Chame test_task_code(task_id=ID, code=código) para validar.\n"
        "   - Se retornar erro: analise a mensagem, corrija o código e teste novamente.\n"
        "   - NUNCA salve um código que falhou no teste.\n"
        "2. Somente após teste bem-sucedido, chame save_task_code(task_id=ID, code=código).\n"
        "3. Retorne o resultado de save_task_code literalmente ao orquestrador.\n\n"
        "REGRA CRÍTICA: o código que o orquestrador envia foi construído com dados reais\n"
        "(endpoints e colunas observadas durante execução interativa). NÃO altere nomes de\n"
        "colunas ou endpoints — corrija apenas erros de sintaxe ou lógica Python.\n\n"
        "## Exemplo de task_code\n"
        "```python\n"
        "def run(from_date, to_date, ctx):\n"
        "    from_d, to_d = ctx.date_range(days=7)\n"
        "    dados = ctx.api(f'/brazil/purchase-orders?from={from_d}&to={to_d}')\n"
        "    df = pd.DataFrame(dados)\n"
        "    fig, ax = plt.subplots(figsize=(10, 4))\n"
        "    ax.bar(df['issue_date'], df['order_number'], color='steelblue')\n"
        "    ax.set_title('Pedidos por Data')\n"
        "    token_chart = ctx.save_chart(fig)\n"
        "    # NÃO use to_markdown() — causa erro de layout no PDF\n"
        "    # Use apenas bullet points com valores agregados\n"
        "    total = df['produced'].sum()\n"
        "    media = df['produced'].mean()\n"
        "    conteudo = (\n"
        "        f'## Produção Diária\\n\\n{token_chart}\\n\\n'\n"
        "        f'- Período: {from_d} a {to_d}\\n'\n"
        "        f'- Total produzido: {total:.0f} unidades\\n'\n"
        "        f'- Média diária: {media:.1f} unidades\\n'\n"
        "    )\n"
        "    return ctx.generate_pdf('Relatório de Produção', conteudo)\n"
        "```\n\n"
        "## Edição de task_code\n"
        "Quando o usuário pedir para alterar qualquer aspecto do relatório (cor, escala, colunas):\n"
        "1. Chame get_task_code para obter o código atual.\n"
        "2. Modifique o código conforme solicitado.\n"
        "3. Chame test_task_code para validar.\n"
        "4. Chame save_task_code para salvar (a versão anterior é arquivada automaticamente).\n\n"
        "## Frequências aceitas\n"
        "once | daily | weekly (requer weekday) | monthly (requer day) |\n"
        "every_Xm (ex: every_2m) | every_Xh (ex: every_6h) | every_Xd (ex: every_3d)\n\n"
        "## Monitores e alertas — frequência e horário\n"
        "Se a instrução recebida mencionar 'monitor', 'alerta', 'avise', 'notifique', 'threshold',\n"
        "'status', 'fique de olho' ou condição a ser verificada periodicamente,\n"
        "use SEMPRE frequency='every_5m' e NÃO inclua time nem weekday nem day.\n"
        "NUNCA pergunte ao usuário que horário deseja — monitores não têm hora fixa.\n"
        "O task_code deve usar ctx.notify() (não ctx.notify_alert()) para disparar notificações.\n\n"
        "## REGRA CRÍTICA — criar vs editar\n"
        "schedule_task SOMENTE quando o usuário pedir explicitamente para CRIAR uma tarefa nova.\n"
        "Para qualquer outra operação sobre tarefa existente (salvar instructions, editar campos,\n"
        "deletar), use a tool correspondente: set_task_instructions, update_scheduled_task ou\n"
        "delete_scheduled_task. NUNCA chame schedule_task quando receber um task_id existente.\n\n"
        "## Salvar instruções\n"
        "Se a instrução recebida começar com 'set_instructions task [ID]:' ou pedir para\n"
        "salvar/definir instructions de uma tarefa existente: extraia o ID e o texto e chame\n"
        "set_task_instructions(task_id=ID, instructions=texto). Nada mais.\n\n"
        "## Ler instruções\n"
        "Se a instrução recebida começar com 'get_instructions task [ID]': chame\n"
        "get_task_instructions(task_id=ID) e retorne o resultado completo.\n\n"
        "## REGRA DE RESPOSTA — OBRIGATÓRIA\n"
        "Sempre retorne o output LITERAL e COMPLETO da tool na sua resposta final.\n"
        "NUNCA resuma, NUNCA substitua por frases como 'concluído' ou 'listagem feita'.\n"
        "O orquestrador depende do conteúdo exato para repassar ao usuário.\n\n"
        "## RACIOCÍNIO OBRIGATÓRIO\n"
        "SEMPRE que for chamar uma tool, inclua na mesma resposta um texto curto explicando "
        "o que vai fazer e por quê, ANTES do bloco de função. "
        "NUNCA emita uma tool call sem texto explicativo."
    )

    llm_sch = llm.bind_tools(sch_tools)
    no_sch_tools = ToolNode(sch_tools)

    def no_scheduler(state: State) -> dict:
        msgs = [SystemMessage(content=SCH_SYSTEM_PROMPT)] + list(state["messages"])
        response = llm_sch.invoke(msgs)
        session = _current_session.get()
        _emit_agent_events(response, lambda ev: _push_event(session, ev))
        return {"messages": [response]}

    def sch_para_onde(state: State) -> str:
        ultima = state["messages"][-1]
        if hasattr(ultima, "tool_calls") and ultima.tool_calls:
            return "tools"
        return END

    builder = StateGraph(State)
    builder.add_node("scheduler", no_scheduler)
    builder.add_node("tools", no_sch_tools)
    builder.add_edge(START, "scheduler")
    builder.add_conditional_edges("scheduler", sch_para_onde, {"tools": "tools", END: END})
    builder.add_edge("tools", "scheduler")

    return builder.compile()


def _build_scheduling_graph(llm):
    """Grafo determinístico para criação e edição de task_code agendado.

    Fluxo create: criar_tarefa → consultar_dados → gerar_codigo → testar_codigo ↔ corrigir_codigo → salvar_codigo → responder
    Fluxo edit:   editar_codigo → testar_codigo ↔ corrigir_codigo → salvar_codigo → responder
    """
    import json as _json
    import re as _re_sg

    _MAX_RETRIES = 3
    _sandbox_skill = (SKILLS_FOLDER / "task_code_sandbox.md").read_text(encoding="utf-8")

    def _extract_code_block(text: str) -> str:
        """Extrai bloco ```task_code``` ou ```python``` que contenha def run."""
        for marker in ("task_code", "python"):
            m = _re_sg.search(rf'```{marker}\n(.*?)```', text, _re_sg.DOTALL)
            if m and "def run" in m.group(1):
                return m.group(1).strip()
        m = _re_sg.search(r'```[^\n]*\n(.*?def run.*?)```', text, _re_sg.DOTALL)
        if m:
            return m.group(1).strip()
        m = _re_sg.search(r'(def run\s*\(from_date.*)', text, _re_sg.DOTALL)
        if m:
            return m.group(1).strip()
        return ""

    def _push(event: dict) -> None:
        _push_event(_current_session.get(), event)

    # ── Node: criar_tarefa ───────────────────────────────────────────────────

    def node_criar_tarefa(state: SchedulingState) -> dict:
        _push({"type": "tool", "name": "schedule_task", "label": "🗓️ Criando registro da tarefa"})

        param_prompt = (
            "Extraia os parâmetros de agendamento da instrução abaixo. "
            "Responda APENAS com JSON válido, sem explicação nem marcação markdown.\n\n"
            f"Instrução: {state['user_request']}\n\n"
            'Formato: {"name": "nome curto (máx 60 chars)", "description": "descrição clara", '
            '"frequency": "once|daily|weekly|monthly|every_Xm|every_Xh|every_Xd", '
            '"weekday": "monday|tuesday|wednesday|thursday|friday|saturday|sunday ou null", '
            '"day": "número 1-31 para monthly ou null", "time": "HH:MM ou null"}\n\n'
            "REGRAS: monitores (me avise/notifique/fique de olho) → frequency=every_5m, time=null. "
            "weekly → weekday obrigatório. monthly → day obrigatório. "
            "Se tempo não informado → time=null."
        )
        response = llm.invoke([HumanMessage(content=param_prompt)])
        try:
            raw = response.content.strip()
            raw = _re_sg.sub(r'^```[^\n]*\n?', '', raw).rstrip('`').strip()
            params = _json.loads(raw)
        except Exception as e:
            return {"error": f"Erro ao interpretar parâmetros de agendamento: {e}", "task_id": "", "task_name": ""}

        call_params: dict = {
            "name": params.get("name", "Nova Tarefa"),
            "description": params.get("description", ""),
            "frequency": params.get("frequency", "daily"),
        }
        if params.get("weekday"):
            call_params["weekday"] = params["weekday"]
        if params.get("day") and params["day"] is not None:
            call_params["day"] = str(params["day"])
        if params.get("time") and params["time"] is not None:
            call_params["time"] = params["time"]

        result = schedule_task.invoke(call_params)

        m = _re_sg.search(r'\[(\d+)\]', result)
        task_id = m.group(1) if m else ""

        if not task_id:
            return {"error": f"Erro ao criar tarefa: {result}", "task_id": "", "task_name": ""}

        return {"task_id": task_id, "task_name": params.get("name", ""), "error": ""}

    # ── Node: consultar_dados ────────────────────────────────────────────────
    # Responsabilidade única: o analista identifica a skill correta, busca dados
    # reais via API e devolve um contexto estruturado (skill, endpoint, colunas,
    # amostra). NÃO escreve código — isso fica no nó seguinte.

    def node_consultar_dados(state: SchedulingState) -> dict:
        if _sub_agent_graph is None:
            return {"error": "Sub-agente analista não inicializado.", "analista_context": ""}

        _push({"type": "tool", "name": "consultar_analista", "label": "🔍 Consultando dados reais"})

        hoje = datetime.now().date().isoformat()

        msg_content = (
            f"[PARA_AGENDAMENTO — FASE 1: COLETA DE CONTEXTO]\n"
            f"[PERÍODO: from={hoje}&to={hoje}]\n\n"
            f"Pedido do usuário: {state['user_request']}\n\n"
            f"Sua tarefa nesta fase:\n"
            f"1. Chame read_skill() para identificar a skill correta para este pedido.\n"
            f"2. Chame chamar_api() para buscar dados reais do período acima.\n"
            f"3. Chame analisar_dataframe() para observar colunas e valores.\n"
            f"4. Retorne um relatório de contexto com EXATAMENTE este formato:\n\n"
            f"SKILL: <nome do arquivo .md usado>\n"
            f"ENDPOINT: <URL completa usada na chamada chamar_api, com parâmetros>\n"
            f"COLUNAS: <lista das colunas relevantes do dataframe>\n"
            f"AMOSTRA: <3-5 linhas representativas do dataframe em formato texto>\n"
            f"TIPO_SAIDA: <PDF_COM_GRAFICO | GRAFICO | EXCEL | MONITOR>\n\n"
            f"⛔ NÃO gere task_code nesta fase.\n"
            f"⛔ NÃO chame gerar_pdf nem gerar_excel.\n"
            f"Apenas colete e reporte o contexto de dados."
        )

        session_id = state["session_id"]
        _sandbox_mode.set(True)
        _sandbox_task_id.set(state.get("task_id", ""))
        try:
            resultado = _sub_agent_graph.invoke(
                {"messages": [HumanMessage(content=msg_content)]},
                config={"configurable": {"thread_id": session_id}, "recursion_limit": 30},
            )
        finally:
            _sandbox_mode.set(False)
            _sandbox_task_id.set("")
        response_text = resultado["messages"][-1].content

        if not response_text or len(response_text.strip()) < 20:
            return {"error": "Analista não retornou contexto de dados.", "analista_context": ""}

        return {"analista_context": response_text, "error": ""}

    # ── Node: gerar_codigo ───────────────────────────────────────────────────
    # Responsabilidade única: dado o contexto do analista, gera o task_code
    # via LLM dedicado com as regras do sandbox como foco total do prompt.

    def node_gerar_codigo(state: SchedulingState) -> dict:
        _push({"type": "tool", "name": "consultar_analista", "label": "⚙️ Gerando código da tarefa"})

        context = state.get("analista_context", "")

        gen_prompt = (
            f"Você é um gerador de task_code Python para tarefas agendadas.\n\n"
            f"## Pedido do usuário\n"
            f"{state['user_request']}\n\n"
            f"## Contexto coletado pelo analista\n"
            f"{context}\n\n"
            f"## Regras obrigatórias do sandbox\n"
            f"{_sandbox_skill}\n\n"
            f"## Instruções\n"
            f"- Escreva a função `def run(from_date, to_date, ctx)` usando o endpoint e colunas acima.\n"
            f"- Use EXATAMENTE o endpoint informado em ENDPOINT (com from_date e to_date como parâmetros de data).\n"
            f"- Use EXATAMENTE os nomes de colunas informados em COLUNAS.\n"
            f"- Siga o template canônico correspondente ao TIPO_SAIDA:\n"
            f"    PDF_COM_GRAFICO → salve gráfico com ctx.save_chart(fig) e gere PDF com ctx.generate_pdf()\n"
            f"    GRAFICO         → retorne apenas ctx.save_chart(fig)\n"
            f"    EXCEL           → retorne ctx.generate_excel(df, 'nome')\n"
            f"    MONITOR         → use ctx.notify() e retorne string com valor atual\n"
            f"- NUNCA use import. pd, np, plt, date, datetime, timedelta etc já estão no namespace.\n"
            f"- Retorne APENAS o bloco Python completo, sem texto explicativo.\n"
            f"- Responda com ```python\\n...\\n```"
        )

        response = llm.invoke([HumanMessage(content=gen_prompt)])
        code = _extract_code_block(response.content)

        if not code:
            return {
                "error": (
                    "LLM não gerou bloco task_code válido. "
                    f"Resposta (primeiros 400 chars): {response.content[:400]}"
                ),
                "task_code": "",
            }

        return {"task_code": code, "error": ""}

    # ── Node: editar_codigo ──────────────────────────────────────────────────

    def node_editar_codigo(state: SchedulingState) -> dict:
        _push({"type": "tool", "name": "get_task_code", "label": "🔍 Lendo código atual da tarefa"})

        result = get_task_code.invoke({"task_id": state["task_id"]})
        current_code = _extract_code_block(result)
        if not current_code and "def run" in result:
            current_code = result

        if not current_code:
            # Sem código existente — gera do zero passando pelos dois nós de criação
            _push({"type": "tool", "name": "consultar_analista", "label": "🔍 Sem código existente — gerando do zero"})
            ctx_result = node_consultar_dados(state)
            if ctx_result.get("error"):
                return ctx_result
            return node_gerar_codigo({**state, **ctx_result})

        _push({"type": "tool", "name": "get_task_code", "label": "✏️ Aplicando modificação ao código"})

        edit_prompt = (
            f"Modifique o código Python abaixo conforme solicitado.\n\n"
            f"MODIFICAÇÃO: {state['user_request']}\n\n"
            f"CÓDIGO ATUAL:\n```python\n{current_code}\n```\n\n"
            f"REGRAS DO SANDBOX:\n{_sandbox_skill}\n\n"
            f"INSTRUÇÕES:\n"
            f"- Faça APENAS a mudança mínima necessária\n"
            f"- Preserve toda a estrutura, variáveis, endpoints e lógica inalterada\n"
            f"- NÃO inclua import — pd, np, plt etc já estão disponíveis\n"
            f"- Retorne APENAS o código Python completo (com def run), sem texto explicativo\n"
            f"- Responda com bloco ```python\\n...\\n```"
        )

        response = llm.invoke([HumanMessage(content=edit_prompt)])
        new_code = _extract_code_block(response.content)
        if not new_code:
            return {"error": f"LLM não retornou código válido: {response.content[:300]}", "task_code": current_code}

        return {"task_code": new_code, "error": ""}

    # ── Node: testar_codigo ──────────────────────────────────────────────────

    def node_testar_codigo(state: SchedulingState) -> dict:
        if not state.get("task_code"):
            return {"error": "Nenhum código para testar.", "retries": state.get("retries", 0)}

        _push({"type": "tool", "name": "test_task_code", "label": "🧪 Testando código"})

        result = test_task_code.invoke({
            "task_id": state["task_id"],
            "code": state["task_code"],
        })

        is_error = result.startswith("❌") or "❌" in result[:30]

        if is_error:
            attempt = state.get("retries", 0) + 1
            try:
                from db import log_task_code_error
                phase = "create" if state.get("mode") == "create" else "edit"
                log_task_code_error(
                    task_id=state["task_id"],
                    attempt=attempt,
                    phase=phase,
                    error=result,
                    code=state["task_code"],
                )
            except Exception:
                pass  # auditoria nunca bloqueia o fluxo principal
            return {
                "error": result,
                "retries": attempt,
            }
        return {"error": "", "result": result}

    # ── Node: corrigir_codigo ────────────────────────────────────────────────

    def node_corrigir_codigo(state: SchedulingState) -> dict:
        _push({"type": "tool", "name": "test_task_code", "label": "🔧 Corrigindo erro no código"})

        error = state["error"]

        # Instrução extra targetada para erros de import bloqueado
        import_match = _re_sg.search(r"Import de '(\w+)' não é permitido", error)
        if not import_match:
            # Também captura "O task_code contém imports, que são bloqueados"
            import_match = _re_sg.search(r"bloqueados no sandbox[:\s]+(.+?)$", error, _re_sg.MULTILINE)
        import_hint = ""
        if import_match:
            blocked = import_match.group(1).strip().split()[0].strip("'\"")
            import_hint = (
                f"\n🚫 ERRO ESPECÍFICO DE IMPORT BLOQUEADO: o módulo '{blocked}' não existe no sandbox.\n"
                f"OBRIGATÓRIO — faça as DUAS coisas:\n"
                f"  1. Remova TODA linha `import {blocked}` (incluindo dentro de def run)\n"
                f"  2. Substitua CADA uso de `{blocked}.X(...)` por equivalente disponível:\n"
                f"     - time.sleep / time.time → desnecessário em tasks, remova\n"
                f"     - time.strftime → use from_date.strftime() ou datetime.now().strftime()\n"
                f"     - math.sqrt/log → use np.sqrt(), np.log()\n"
                f"     - json.dumps/loads → use str() ou dict diretamente\n"
                f"     - re.match/search → reescreva sem regex se possível\n"
                f"     - Para qualquer outro: verifique se pd, np, datetime, timedelta cobrem a necessidade\n"
            )

        fix_prompt = (
            f"O código Python abaixo falhou no teste. Corrija-o.\n\n"
            f"ERRO:\n{error}\n"
            f"{import_hint}\n"
            f"CÓDIGO COM ERRO:\n```python\n{state['task_code']}\n```\n\n"
            f"REGRAS DO SANDBOX:\n{_sandbox_skill}\n\n"
            f"INSTRUÇÕES:\n"
            f"- Corrija APENAS o erro acima, preservando toda a lógica restante\n"
            f"- NÃO inclua NENHUM import — pd, np, plt, date, datetime, timedelta etc já estão disponíveis\n"
            f"- Retorne APENAS o código Python completo (com def run), sem texto explicativo\n"
            f"- Responda com bloco ```python\\n...\\n```"
        )

        response = llm.invoke([HumanMessage(content=fix_prompt)])
        new_code = _extract_code_block(response.content)
        if not new_code:
            return {}  # keep existing task_code, will fail again on next test

        # Pós-processamento determinístico: remove linhas de import que o LLM teima em gerar
        stripped_lines = [
            ln for ln in new_code.splitlines()
            if not _re_sg.match(r'\s*(import |from \S+ import )', ln)
        ]
        new_code = "\n".join(stripped_lines)

        return {"task_code": new_code}

    # ── Node: salvar_codigo ──────────────────────────────────────────────────

    def node_salvar_codigo(state: SchedulingState) -> dict:
        _push({"type": "tool", "name": "save_task_code", "label": "💾 Salvando código validado"})

        result = save_task_code.invoke({
            "task_id": state["task_id"],
            "code": state["task_code"],
        })

        return {"result": result}

    # ── Node: responder ──────────────────────────────────────────────────────

    def node_responder(state: SchedulingState) -> dict:
        tid = state.get("task_id", "?")
        err = state.get("error", "")
        code = state.get("task_code", "")

        if err and not code:
            msg = f"Não foi possível completar a operação: {err}"
        elif err:
            retries = state.get("retries", 0)
            # Tarefa foi criada como draft mas o código falhou — remove o registro
            if state.get("mode") == "create" and tid and tid != "?":
                try:
                    from db import get_db as _get_db
                    with _get_db() as _conn:
                        _conn.execute("DELETE FROM scheduled_tasks WHERE id = ? AND status = 'draft'", (tid,))
                        _conn.commit()
                except Exception:
                    pass
            msg = (
                f"Não foi possível gerar o código da tarefa após {retries} tentativa(s).\n\n"
                f"**Último erro:**\n```\n{err[:500]}\n```\n\n"
                f"Tente novamente ou reformule o pedido com mais detalhes."
            )
        elif state["mode"] == "create":
            msg = (
                f"Tarefa **[{tid}]** criada e agendada com código determinístico.\n\n"
                f"Para ajustar: **editar tarefa {tid}**"
            )
        else:
            msg = (
                f"Tarefa **[{tid}]** atualizada com sucesso.\n\n"
                f"Para novos ajustes: **editar tarefa {tid}**"
            )

        return {"result": msg}

    # ── Routing ──────────────────────────────────────────────────────────────

    def route_entry(state: SchedulingState) -> str:
        return "criar_tarefa" if state["mode"] == "create" else "editar_codigo"

    def route_after_criar(state: SchedulingState) -> str:
        if not state.get("task_id") or state.get("error"):
            return "responder"
        return "consultar_dados"

    def route_after_test(state: SchedulingState) -> str:
        if not state.get("error"):
            return "salvar_codigo"
        if state.get("retries", 0) < _MAX_RETRIES:
            return "corrigir_codigo"
        return "responder"

    # ── Build ────────────────────────────────────────────────────────────────

    def route_after_consultar(state: SchedulingState) -> str:
        if state.get("error"):
            return "responder"
        return "gerar_codigo"

    builder = StateGraph(SchedulingState)
    builder.add_node("criar_tarefa",    node_criar_tarefa)
    builder.add_node("consultar_dados", node_consultar_dados)
    builder.add_node("gerar_codigo",    node_gerar_codigo)
    builder.add_node("editar_codigo",   node_editar_codigo)
    builder.add_node("testar_codigo",   node_testar_codigo)
    builder.add_node("corrigir_codigo", node_corrigir_codigo)
    builder.add_node("salvar_codigo",   node_salvar_codigo)
    builder.add_node("responder",       node_responder)

    builder.add_conditional_edges(START, route_entry, {
        "criar_tarefa": "criar_tarefa",
        "editar_codigo": "editar_codigo",
    })
    builder.add_conditional_edges("criar_tarefa", route_after_criar, {
        "consultar_dados": "consultar_dados",
        "responder": "responder",
    })
    builder.add_conditional_edges("consultar_dados", route_after_consultar, {
        "gerar_codigo": "gerar_codigo",
        "responder": "responder",
    })
    builder.add_edge("gerar_codigo",  "testar_codigo")
    builder.add_edge("editar_codigo", "testar_codigo")
    builder.add_conditional_edges("testar_codigo", route_after_test, {
        "salvar_codigo":  "salvar_codigo",
        "corrigir_codigo": "corrigir_codigo",
        "responder":      "responder",
    })
    builder.add_edge("corrigir_codigo", "testar_codigo")
    builder.add_edge("salvar_codigo",   "responder")
    builder.add_edge("responder",       END)

    return builder.compile()


@tool
def gerenciar_agenda(instrucao: str) -> str:
    """Delega ao sub-agente de scheduling operações sobre tarefas agendadas.

    Use para: criar, listar, editar, deletar tarefas e salvar instruções de execução.

    Exemplos de instruções:
      - "Criar tarefa 'Relatório OEE Semanal', frequência weekly, weekday monday, time 08:00"
      - "Listar todas as tarefas agendadas"
      - "Deletar tarefa 003"
      - "set_instructions task 001: Passo 1 — buscar produção dos últimos 7 dias com
         consultar_analista(periodo='last_7_days'). Passo 2 — gerar PDF com gerar_pdf()."

    Args:
        instrucao: Descrição completa da operação. Para salvar instruções de uma tarefa
                   recém-criada, use o formato 'set_instructions task [ID]: [passos]'.
    """
    if _scheduler_agent_graph is None:
        return "Sub-agente de scheduling não inicializado."
    resultado = _scheduler_agent_graph.invoke(
        {"messages": [HumanMessage(content=instrucao)]},
        config={"configurable": {"thread_id": _current_session.get()}, "recursion_limit": 15},
    )
    return resultado["messages"][-1].content


@tool
def criar_tarefa_agendada(instrucao: str) -> str:
    """Cria uma nova tarefa agendada com código Python determinístico gerado automaticamente.

    Use para CRIAR/AGENDAR uma nova tarefa automática, relatório periódico ou monitor.
    Este tool executa o fluxo completo:
      1. Infere parâmetros (nome, frequência, horário) da instrução
      2. Cria o registro da tarefa no banco
      3. Executa análise prévia para observar os dados reais
      4. Gera o código Python (def run) baseado nos dados reais
      5. Testa o código (até 3 tentativas de correção automática)
      6. Salva o código validado

    Args:
        instrucao: Descrição completa do que a tarefa deve fazer, com frequência e
                   qualquer parâmetro relevante. Exemplos:
                   "Relatório semanal de OEE toda segunda às 8h"
                   "Monitor de produção a cada 5 minutos, alerta se OEE < 80%"
                   "Planilha mensal de defeitos no dia 1 de cada mês"
    """
    if _scheduling_graph is None:
        return "SchedulingGraph não inicializado."

    session_id = _current_session.get()
    _push_event(session_id, {"type": "tool", "name": "criar_tarefa_agendada", "label": "🗓️ Iniciando criação de tarefa agendada"})

    initial_state: SchedulingState = {
        "mode": "create",
        "session_id": session_id,
        "user_request": instrucao,
        "task_id": "",
        "task_name": "",
        "task_code": "",
        "error": "",
        "retries": 0,
        "result": "",
    }

    resultado = _scheduling_graph.invoke(initial_state)  # type: ignore[arg-type]
    return resultado.get("result", "Tarefa criada.")


@tool
def editar_tarefa_agendada(task_id: str, modificacao: str) -> str:
    """Edita o código Python de uma tarefa agendada existente.

    Use quando o usuário pedir para EDITAR/ALTERAR/AJUSTAR qualquer aspecto de uma
    tarefa existente: cor, escala, colunas, períodos, threshold, lógica, etc.
    Este tool executa o fluxo completo:
      1. Lê o código atual da tarefa
      2. Aplica a modificação via LLM
      3. Testa o código modificado (até 3 tentativas de correção automática)
      4. Salva o código validado

    Args:
        task_id: ID da tarefa a editar (ex: "001", "42").
        modificacao: O que deve ser alterado. Exemplos:
                     "Mude a cor das barras para laranja"
                     "Adicione linha de meta (target) no gráfico"
                     "Mude o threshold de OEE de 80% para 75%"
    """
    if _scheduling_graph is None:
        return "SchedulingGraph não inicializado."

    session_id = _current_session.get()
    _push_event(session_id, {"type": "tool", "name": "editar_tarefa_agendada", "label": "✏️ Editando código da tarefa"})

    initial_state: SchedulingState = {
        "mode": "edit",
        "session_id": session_id,
        "user_request": modificacao,
        "task_id": task_id,
        "task_name": "",
        "task_code": "",
        "error": "",
        "retries": 0,
        "result": "",
    }

    resultado = _scheduling_graph.invoke(initial_state)  # type: ignore[arg-type]
    return resultado.get("result", "Tarefa atualizada.")


# ── Orquestrador ──────────────────────────────────────────────────────────────

MAX_INTERACOES = int(os.getenv("MAX_INTERACOES", "10"))


def _build_orchestrator(llm, checkpointer=None):
    ORQ_SYSTEM_PROMPT = (
        "Você é o agente orquestrador de um sistema de monitoramento industrial.\n\n"
        "## HISTÓRICO DE CONVERSA\n"
        "As mensagens acima (antes desta instrução) são o histórico completo desta sessão, "
        "incluindo perguntas e respostas anteriores ao último refresh da página do usuário. "
        "NUNCA diga que não tem memória de conversas anteriores — você tem acesso total ao "
        "histórico desta sessão. Se o usuário perguntar o que perguntou antes, o que foi "
        "analisado, ou pedir para repetir algo, consulte as mensagens anteriores.\n\n"
        "## ROTEAMENTO — leia isto antes de qualquer ação\n"
        "Classifique o pedido do usuário em UMA das categorias abaixo e siga APENAS o fluxo correspondente:\n\n"
        "  A) CONSULTA PONTUAL — o usuário quer ver dados, gráficos, tabelas ou análises AGORA.\n"
        "     Palavras-chave: 'me diga', 'mostre', 'qual é', 'quanto', 'como está', 'gráfico de',\n"
        "     'produção de hoje', 'análise de', 'compare', 'relatório de'.\n"
        "     → Chame calcular_periodo() + consultar_analista(). NÃO envolva gerenciar_agenda.\n\n"
        "  B) PERGUNTA CONCEITUAL — o usuário quer saber o que algo significa ou como é calculado.\n"
        "     → Chame a tool rag_* mais adequada ao tema da pergunta. Responda diretamente sem buscar dados.\n"
        "       rag_dominio()      → KPIs, painéis, interpretação, linhas, turnos\n"
        "       rag_capacidades()  → o que o agente pode fazer, agendamento, monitores\n"
        "       rag_arquitetura()  → como o sistema funciona, sub-agentes, skills\n"
        "       rag_dados()        → endpoints da API, schema do banco, colunas\n\n"
        "  C) AGENDAMENTO — o usuário quer CRIAR, EDITAR, LISTAR, PAUSAR ou DELETAR uma tarefa recorrente.\n"
        "     Palavras-chave obrigatórias: 'agendar', 'todo dia', 'toda semana', 'automaticamente',\n"
        "     'criar tarefa', 'monitor', 'me avise quando', 'editar tarefa [ID]'.\n"
        "     → Siga o fluxo de agendamento detalhado abaixo.\n\n"
        "  D) PERGUNTA SIMPLES — data, hora, saudação, dúvida geral sem dados.\n"
        "     → Responda diretamente sem chamar nenhuma tool de dados.\n\n"
        "  E) PERGUNTA SOBRE O DASHBOARD (texto ou imagem) — o usuário pergunta sobre o que está\n"
        "     visível na tela: valores nos gráficos, KPIs, alertas, notificações, comparações entre\n"
        "     linhas/turnos. Palavras-chave: 'na tela', 'no dashboard', 'o que aparece', 'o gráfico\n"
        "     mostra', 'quantos alertas', 'quantas notificações', 'KPI'. Também se enquadra aqui\n"
        "     qualquer mensagem que contenha uma imagem anexada pelo usuário.\n"
        "     → Chame get_dashboard_charts() para obter os dados numéricos do dashboard.\n"
        "     → Se a pergunta envolver o SIGNIFICADO de uma métrica (o que é FPY, como se calcula\n"
        "       OEE, o que representa downtime), chame rag() em seguida para complementar.\n"
        "     → NÃO chame consultar_analista — os dados relevantes já estão na tela.\n\n"
        "  F) PAINEL CUSTOMIZADO NO DASHBOARD — o usuário quer ADICIONAR, INSERIR, FIXAR ou REMOVER\n"
        "     um gráfico como painel permanente no dashboard (seção 'Painéis Customizados').\n"
        "     Palavras-chave OBRIGATÓRIAS: 'adiciona ao dashboard', 'insere no dashboard',\n"
        "     'quero ver no dashboard', 'cria um painel com', 'fixar no dashboard',\n"
        "     'coloca no dashboard', 'remove o painel', 'lista os painéis'.\n"
        "     ⚠️ NÃO confunda com pedidos de gráfico no chat (categoria A).\n"
        "     Um gráfico no chat é temporário. Um painel no dashboard é permanente.\n"
        "     → Siga o fluxo de painel customizado detalhado abaixo.\n\n"
        "REGRA CRÍTICA: na dúvida entre A e F, escolha A. NUNCA adicione um painel ao dashboard\n"
        "sem que o usuário tenha usado explicitamente palavras de inserção no dashboard (categoria F).\n\n"
        "## Análise de imagens do dashboard\n"
        "Quando o usuário envia uma imagem junto com sua mensagem, trata-se de um recorte da\n"
        "própria tela do dashboard — pode conter gráficos de produção, FPY, OEE, defeitos,\n"
        "KPIs, alertas ou qualquer outro elemento visual do sistema de monitoramento industrial.\n"
        "Analise visualmente o conteúdo da imagem e:\n"
        "  1. Descreva o que está sendo exibido (tipo de gráfico, eixos, tendências, valores\n"
        "     destacados, anomalias visíveis).\n"
        "  2. Se precisar dos valores numéricos exatos por trás do gráfico, chame\n"
        "     get_dashboard_charts() — ela retorna os dados brutos de todos os gráficos.\n"
        "  3. Se precisar explicar o significado de uma métrica visível (FPY, OEE, downtime,\n"
        "     etc.), chame rag() para buscar a definição precisa do sistema.\n"
        "  4. Combine o que você vê na imagem com os dados retornados pelas tools para dar\n"
        "     uma resposta completa e contextualizada.\n\n"
        "## Follow-up sobre gráficos gerados por consultar_analista\n"
        "Quando o usuário fizer uma pergunta de follow-up sobre um gráfico gerado nesta sessão,\n"
        "consulte o histórico: a resposta do sub-agente inclui um bloco **Dados do gráfico:**\n"
        "com os pontos-chave (totais, máximos, mínimos, médias). Use esses dados para responder\n"
        "diretamente — SEM chamar consultar_analista novamente.\n"
        "Se a pergunta for visual ('por que há um pico?', 'onde está a queda?', 'o que mostra\n"
        "o trecho X?') ou se os dados textuais do histórico forem insuficientes, chame\n"
        "ver_grafico(chart_id=UUID) — ela injeta a imagem no contexto para análise visual.\n"
        "Se o usuário pedir uma MODIFICAÇÃO VISUAL do gráfico (cor, paleta, estilo, tipo de "
        "gráfico, labels, título, etc.), chame consultar_analista novamente passando em `detalhes` "
        "o pedido original completo E a modificação solicitada. Exemplo: 'Gráfico de produção "
        "diária agregada por dia — mesmos dados e período, mas com paleta de cores cinza/preto.'\n"
        "O UUID está no histórico no formato [chart:UUID].\n\n"
        "REGRA CRÍTICA: na dúvida entre A e C, escolha A. Nunca inicie um fluxo de agendamento\n"
        "a menos que o usuário tenha usado explicitamente palavras de agendamento (categoria C).\n\n"
        "## Agendamento de tarefas — CRIAR nova tarefa\n"
        "Quando o usuário pedir para criar uma tarefa agendada (relatório periódico, monitor,\n"
        "alerta automático), chame APENAS:\n\n"
        "  criar_tarefa_agendada(instrucao='descrição completa do que a tarefa deve fazer,\n"
        "                                   frequência e qualquer parâmetro relevante')\n\n"
        "O fluxo completo (criar registro, executar análise, gerar código, testar, salvar)\n"
        "é executado automaticamente dentro do tool. NUNCA faça esse fluxo manualmente.\n"
        "NUNCA peça ao usuário nome ou descrição — infira do contexto.\n"
        "Inclua na instrução: o que a tarefa faz, a frequência e (se mencionado) o horário.\n"
        "Exemplos de instrucao:\n"
        "  'Relatório semanal de OEE toda segunda às 8h'\n"
        "  'Monitor de OEE a cada 5 minutos, alerta se OEE < 80%'\n"
        "  'Planilha mensal de defeitos no dia 1 de cada mês'\n\n"
        "Após criar_tarefa_agendada retornar, repasse o resultado ao usuário.\n"
        "NÃO inclua tokens [chart:uuid], [pdf:uuid] ou [excel:uuid] na sua resposta —\n"
        "os artefatos já foram enviados ao chat automaticamente.\n\n"
        "## Agendamento de tarefas — EDITAR tarefa existente\n"
        "Quando o usuário pedir para editar/ajustar qualquer aspecto de uma tarefa\n"
        "(cor, escala, colunas, threshold, período, lógica), chame APENAS:\n\n"
        "  editar_tarefa_agendada(task_id='ID', modificacao='o que deve mudar')\n\n"
        "O fluxo completo (ler código atual, aplicar mudança, testar, salvar) é automático.\n"
        "NUNCA leia, modifique ou salve o código manualmente para edição.\n"
        "NÃO inclua tokens na sua resposta após a edição — os artefatos já foram enviados.\n\n"
        "## Agendamento de tarefas — operações simples\n"
        "Para LISTAR, PAUSAR, RETOMAR ou DELETAR tarefas: chame gerenciar_agenda(instrucao).\n"
        "Para VER o código atual de uma tarefa: chame get_task_code(task_id=ID).\n\n"
        "## Painéis customizados no dashboard — fluxo obrigatório ao CRIAR\n"
        "Quando o usuário pedir para ADICIONAR/INSERIR/FIXAR um gráfico no dashboard,\n"
        "siga EXATAMENTE estes passos:\n\n"
        "1. Confirme a intenção antes de agir:\n"
        "   Responda: 'Vou criar e fixar um painel [descrição] no dashboard. Confirma?'\n"
        "   Aguarde confirmação. Se o usuário confirmar, prossiga.\n\n"
        "2. Execute consultar_analista() com para_agendamento=True para descobrir\n"
        "   os dados reais (endpoint, colunas, valores). Chame calcular_periodo() antes.\n"
        "   ⛔ OBRIGATÓRIO — nunca escreva código sem ter observado os dados reais.\n\n"
        "3. Escreva o widget_code com base no bloco task_context do passo 2.\n"
        "   REGRAS CRÍTICAS DO WIDGET CODE (viole qualquer uma = erro garantido):\n\n"
        "   ⛔ NUNCA use `import` dentro de run() — o sandbox bloqueia qualquer import.\n"
        "      Variáveis pré-injetadas disponíveis: pd, np, ctx, from_date, to_date,\n"
        "      date, datetime, timedelta. Use-as diretamente.\n\n"
        "   ⛔ ctx.api() SEMPRE retorna lista Python (list[dict]). NUNCA é um DataFrame.\n"
        "      Converta SEMPRE: `df = pd.DataFrame(ctx.api('/endpoint...'))`\n"
        "      Acessar colunas antes de converter causa KeyError.\n\n"
        "   ⛔ run() DEVE retornar um dict Chart.js com 'type' e 'data'.\n"
        "      Tipos aceitos: bar, line, pie, doughnut, radar, polarArea.\n"
        "      'data' deve ter: 'labels' (list[str]) e 'datasets' (list[dict]).\n"
        "      Cada dataset: {'label': str, 'data': list[float]}.\n\n"
        "   ⛔ NUNCA escreva datas específicas — use from_date/to_date recebidos como parâmetro.\n\n"
        "   Template COMPLETO e correto (copie e adapte):\n"
        "   ```python\n"
        "   def run(from_date, to_date, ctx):\n"
        "       raw = ctx.api(f'/brazil/order-items/summary?from={from_date}&to={to_date}')\n"
        "       df = pd.DataFrame(raw['by_status'])  # OBRIGATÓRIO: converter list→DataFrame\n"
        "       labels = df['status'].tolist()\n"
        "       values = df['total_items'].tolist()\n"
        "       return {\n"
        "           'type': 'pie',\n"
        "           'data': {\n"
        "               'labels': labels,\n"
        "               'datasets': [{'label': 'Itens por Status', 'data': values}]\n"
        "           }\n"
        "       }\n"
        "   ```\n\n"
        "4. Chame test_widget_code(title=TÍTULO, code=CÓDIGO) para validar.\n"
        "   Se retornar erro, leia a mensagem, corrija O ERRO EXATO e teste novamente.\n"
        "   Erros comuns: (a) import bloqueado → remova o import, use a variável pré-injetada;\n"
        "   (b) KeyError/AttributeError → você esqueceu o pd.DataFrame() após ctx.api().\n"
        "   NUNCA avance com erro.\n\n"
        "5. Chame add_chart_to_dashboard(title=TÍTULO, description=DESCRIÇÃO, code=CÓDIGO).\n\n"
        "6. Responda: 'Painel **[título]** adicionado ao dashboard. "
        "Ele aparece na seção Painéis Customizados ao final da página.'\n\n"
        "## Painéis customizados — listar, editar e remover\n"
        "- Para LISTAR painéis: chame list_dashboard_widgets().\n"
        "- Para LER o código de um painel: chame get_widget_code(widget_id=ID).\n"
        "- Para EDITAR um painel existente (alterar cores, lógica, título etc.):\n"
        "  1. list_dashboard_widgets → obter o ID\n"
        "  2. get_widget_code → ler o código atual\n"
        "  3. Modificar o código conforme solicitado\n"
        "  4. test_widget_code → validar\n"
        "  5. update_widget(widget_id, code, title?, description?) → salvar\n"
        "  NÃO delete e recrie — use update_widget para editar in-place.\n"
        "- Para REMOVER um painel: chame delete_dashboard_widget(widget_id=ID).\n\n"
        "## RACIOCÍNIO OBRIGATÓRIO\n"
        "SEMPRE que for chamar uma tool, você DEVE incluir na mesma resposta um texto curto "
        "em linguagem natural explicando o que vai fazer e por quê, ANTES do bloco de função. "
        "Exemplo: 'Vou calcular o período solicitado para obter as datas exatas.' "
        "O texto e a chamada de tool devem vir juntos na mesma resposta. "
        "NUNCA emita uma tool call sem texto explicativo."
    )

    orq_tools = [
        get_current_datetime, get_dashboard_charts, calcular_periodo, consultar_analista,
        ver_grafico,
        rag_dominio, rag_capacidades, rag_arquitetura, rag_dados,
        gerar_pdf, gerar_excel,
        gerenciar_agenda,
        criar_tarefa_agendada, editar_tarefa_agendada,
        get_task_code,
        test_widget_code, add_chart_to_dashboard,
        list_dashboard_widgets, delete_dashboard_widget,
        get_widget_code, update_widget,
    ]
    llm_orq = llm.bind_tools(orq_tools)
    no_orq_tools = ToolNode(orq_tools)

    def no_orquestrador(state: State) -> dict:
        historico_recente = state["messages"][-(MAX_INTERACOES * 2):]
        msgs = [SystemMessage(content=ORQ_SYSTEM_PROMPT)] + historico_recente
        return {"messages": [llm_orq.invoke(msgs)]}

    def orq_para_onde(state: State) -> str:
        ultima = state["messages"][-1]
        if hasattr(ultima, "tool_calls") and ultima.tool_calls:
            return "tools"
        return END

    builder = StateGraph(State)
    builder.add_node("orquestrador", no_orquestrador)
    builder.add_node("tools", no_orq_tools)
    builder.add_edge(START, "orquestrador")
    builder.add_conditional_edges("orquestrador", orq_para_onde, {"tools": "tools", END: END})
    builder.add_edge("tools", "orquestrador")

    return builder.compile(checkpointer=checkpointer)


# ── API pública ───────────────────────────────────────────────────────────────

def init_multi_agent(project: str, location: str, model_name: str, credentials=None) -> None:
    """Inicializa orquestrador e sub-agente. Chamado no startup do FastAPI."""
    global _orchestrator_graph, _sub_agent_graph, _scheduler_agent_graph, _scheduling_graph, _checkpointer
    try:
        from langchain_google_vertexai import ChatVertexAI
        from langgraph.checkpoint.sqlite import SqliteSaver

        llm = ChatVertexAI(
            model_name=model_name,
            project=project,
            location=location,
            credentials=credentials,
            temperature=0.7,
        )

        # Conexão persistente para o checkpointer (check_same_thread=False pois
        # o FastAPI despacha requests em threads diferentes)
        conn = sqlite3.connect(str(DB_PATH), check_same_thread=False)
        _checkpointer = SqliteSaver(conn)

        chart_store.init_chart_store(DB_PATH)
        _sub_agent_graph = _build_sub_agent(llm)
        _scheduler_agent_graph = _build_scheduler_agent(llm)
        _scheduling_graph = _build_scheduling_graph(llm)
        _orchestrator_graph = _build_orchestrator(llm, _checkpointer)
        logger.info("Multi-agente inicializado: orquestrador + analista + scheduler + scheduling_graph")
    except Exception:
        import traceback
        logger.warning("Falha ao inicializar multi-agente:\n%s", traceback.format_exc())


def _build_human_message(query: str, session_id: str) -> HumanMessage:
    """Constrói HumanMessage simples ou multimodal se houver imagens no snapshot."""
    with _cs_lock:
        snapshot = _chart_snapshots.get(session_id, {})
    images = snapshot.get("images_b64") or []
    if not images:
        return HumanMessage(content=query)
    parts: list = [{"type": "text", "text": query or "Analise as imagens anexadas."}]
    for b64 in images:
        parts.append({"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}})
    return HumanMessage(content=parts)


def invoke_multi_agent(query: str, session_id: str = "default") -> str:
    """Executa o orquestrador com a query do usuário e retorna a resposta final.

    O session_id isola o histórico de conversa — cada sessão tem seu próprio contexto
    persistido no mfg.db via SqliteSaver, e seus DataFrames no store em memória.
    """
    if _orchestrator_graph is None:
        raise RuntimeError("Multi-agente não inicializado.")
    _current_session.set(session_id)
    resultado = _orchestrator_graph.invoke(
        {"messages": [_build_human_message(query, session_id)]},
        config={"configurable": {"thread_id": session_id}, "recursion_limit": 50},
    )
    return resultado["messages"][-1].content



def stream_multi_agent(query: str, session_id: str = "default"):
    """Gerador que yields dicts de eventos SSE enquanto o agente processa.

    Tipos de evento:
      {"type": "thinking", "text": "..."}   — bloco de raciocínio do modelo
      {"type": "tool",     "name": "...", "label": "..."}  — tool sendo chamada
      {"type": "reply",    "text": "..."}   — resposta final (markdown)
      {"type": "error",    "text": "..."}   — erro irrecuperável

    Arquitetura de threading:
      - O orquestrador roda em uma thread separada e coloca eventos numa queue.Queue.
      - O sub-agente (dentro do ToolNode) também coloca eventos via _push_event na mesma
        queue. Como rodam em threads diferentes, os eventos chegam em tempo real ao gerador.
    """
    if _orchestrator_graph is None:
        yield {"type": "error", "text": "Agente não inicializado."}
        return

    _current_session.set(session_id)

    event_queue: _queue_module.Queue = _queue_module.Queue()
    with _eq_lock:
        _event_queues[session_id] = event_queue

    config = {"configurable": {"thread_id": session_id}, "recursion_limit": 50}

    human_message = _build_human_message(query, session_id)

    def _run_orchestrator():
        # ContextVar não se propaga para threading.Thread — precisa setar explicitamente.
        _current_session.set(session_id)
        try:
            for chunk in _orchestrator_graph.stream(
                {"messages": [human_message]},
                config=config,
                stream_mode="updates",
            ):
                for node_name, update in chunk.items():
                    messages = update.get("messages", [])
                    if not isinstance(messages, list):
                        messages = [messages]

                    for msg in messages:
                        _emit_agent_events(msg, event_queue.put)

                        if node_name == "orquestrador":
                            tool_calls = getattr(msg, "tool_calls", None) or []
                            fc = getattr(msg, "additional_kwargs", {}).get("function_call")
                            has_tool = bool(tool_calls or fc)
                            text = _extract_text_content(msg)
                            if text and not has_tool:
                                event_queue.put({"type": "reply", "text": text})

        except Exception as e:
            import traceback as _tb
            import sys
            print(f"\n[STREAM ERROR] {e}\n{_tb.format_exc()}", file=sys.stderr, flush=True)
            logger.error("[STREAM ERROR] %s\n%s", e, _tb.format_exc())
            event_queue.put({"type": "error", "text": f"Erro interno: {e}"})
        finally:
            event_queue.put(None)  # sentinel — sinaliza fim do stream

    orchestrator_thread = threading.Thread(target=_run_orchestrator, daemon=True)
    orchestrator_thread.start()

    try:
        while True:
            event = event_queue.get()
            if event is None:
                break
            yield event
    finally:
        with _eq_lock:
            _event_queues.pop(session_id, None)


def is_multi_agent_ready() -> bool:
    return _orchestrator_graph is not None
